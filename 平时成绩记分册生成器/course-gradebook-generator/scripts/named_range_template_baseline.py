"""Build the current-environment protected baseline for gradebook v1.1."""

from __future__ import annotations

import math
import shutil
import struct
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import olefile
from openpyxl import load_workbook

from named_range_contracts import required_names
from named_range_utils import (
    compare_named_range_inventories,
    expected_named_range_locations,
    set_named_range_from_location,
    validate_named_range_inventory,
)
from package_common import V10_MANIFEST, V10_TEMPLATE, V11_MANIFEST, load_manifest
from validate_template import convert_to_xls, convert_to_xlsx
from xls_named_range_utils import (
    compare_xls_and_xlsx_named_ranges,
    normalize_libreoffice_print_title_records,
    normalize_xls_summary_information,
    validate_xls_named_range_inventory,
)


@dataclass(frozen=True)
class ControlledV11Baseline:
    """Artifacts produced by the same controlled v1.1 construction pipeline."""

    controlled_v11_xls: Path
    controlled_v11_xlsx: Path
    controlled_workbook: Any
    raw_xls_inventory: dict[str, Any]
    xlsx_inventory: dict[str, Any]
    expected_locations: dict[str, Any]


@dataclass(frozen=True)
class ControlledV11Roundtrip:
    """A committed v1.1 candidate after the same controlled conversion path."""

    controlled_xls: Path
    controlled_xlsx: Path
    controlled_workbook: Any
    raw_xls_inventory: dict[str, Any]
    xlsx_inventory: dict[str, Any]


_BASELINE_CACHE: dict[tuple[str, str], ControlledV11Baseline] = {}
_CANDIDATE_CACHE: dict[tuple[str, str, str], ControlledV11Roundtrip] = {}


def _biff_records(data: bytes):
    position = 0
    while position + 4 <= len(data):
        record_type, record_length = struct.unpack_from("<HH", data, position)
        end = position + 4 + record_length
        if end > len(data):
            raise RuntimeError(
                f"BIFF record is truncated at offset {position}: type={record_type} length={record_length}"
            )
        yield position, record_type, data[position:end]
        position = end
    if position != len(data):
        raise RuntimeError(f"BIFF stream has {len(data) - position} trailing bytes")


def _biff_name_from_record(record: bytes) -> str:
    payload = record[4:]
    if len(payload) < 15:
        raise RuntimeError("BIFF NAME record is truncated before its name")
    name_length = int(payload[3])
    name_end = 15 + name_length
    if name_end > len(payload):
        raise RuntimeError("BIFF NAME record has a truncated name")
    return payload[15:name_end].decode("latin1")


def _regular_sector_chain(start_sector: int, fat: list[int]) -> list[int]:
    chain: list[int] = []
    sector = start_sector
    while sector not in (olefile.ENDOFCHAIN, olefile.FREESECT):
        if sector < 0 or sector >= len(fat) or sector in chain:
            raise RuntimeError(f"Invalid OLE sector chain at sector {sector}")
        chain.append(sector)
        sector = fat[sector]
    if sector != olefile.ENDOFCHAIN:
        raise RuntimeError("OLE stream chain ended at a free sector")
    return chain


def _fat_sector_ids(file_bytes: bytearray, sector_size: int) -> list[int]:
    fat_sector_count = struct.unpack_from("<I", file_bytes, 44)[0]
    result: list[int] = []
    for index in range(109):
        sector = struct.unpack_from("<I", file_bytes, 76 + index * 4)[0]
        if sector != olefile.FREESECT:
            result.append(sector)
    if len(result) != fat_sector_count:
        raise RuntimeError(
            f"OLE DIFAT lists {len(result)} FAT sectors, header declares {fat_sector_count}"
        )
    if not result:
        raise RuntimeError("OLE file has no FAT sector")
    return result


def materialize_stable_v11_xls(
    canonical_v10_xls: Path,
    controlled_v11_xls: Path,
    output_xls: Path,
) -> Path:
    """Add managed names while preserving canonical v1.0 BIFF formatting records.

    LibreOffice's XLS exporter quantizes some column and page properties on
    each platform. The committed v1.1 package therefore uses the v1.0
    workbook stream as its protected source and imports only the validated
    managed NAME records produced by the controlled builder. BoundSheet
    offsets and the compound-file stream chain are updated explicitly.
    """
    canonical_v10_xls = Path(canonical_v10_xls).expanduser().resolve()
    controlled_v11_xls = Path(controlled_v11_xls).expanduser().resolve()
    output_xls = Path(output_xls).expanduser().resolve()
    if output_xls in (canonical_v10_xls, controlled_v11_xls):
        raise RuntimeError("Stable v1.1 output must be a new file")
    with olefile.OleFileIO(str(canonical_v10_xls)) as canonical_ole:
        canonical_stream = canonical_ole.openstream("Workbook").read()
        canonical_fat = list(canonical_ole.fat)
        canonical_entry = next(
            entry for entry in canonical_ole.direntries if entry and entry.name == "Workbook"
        )
        canonical_chain = _regular_sector_chain(canonical_entry.isectStart, canonical_fat)
        canonical_dir_chain = _regular_sector_chain(canonical_ole.first_dir_sector, canonical_fat)
        sector_size = int(canonical_ole.sector_size)
    with olefile.OleFileIO(str(controlled_v11_xls)) as controlled_ole:
        controlled_stream = controlled_ole.openstream("Workbook").read()

    managed_records = [
        record
        for _, record_type, record in _biff_records(controlled_stream)
        if record_type == 0x0018 and _biff_name_from_record(record).startswith("gb_")
    ]
    expected_names = set(required_names("with_skill"))
    actual_names = {_biff_name_from_record(record) for record in managed_records}
    if actual_names != expected_names or len(managed_records) != len(expected_names):
        raise RuntimeError(
            "Controlled v1.1 NAME records do not match the managed contract: "
            f"expected={sorted(expected_names)} actual={sorted(actual_names)}"
        )

    inserted_records = b"".join(managed_records)
    canonical_stream_mutable = bytearray(canonical_stream)
    first_global_eof = next(
        position for position, record_type, _ in _biff_records(canonical_stream) if record_type == 0x000A
    )
    for position, record_type, _ in _biff_records(canonical_stream):
        if record_type != 0x0085 or position >= first_global_eof:
            continue
        sheet_offset = struct.unpack_from("<I", canonical_stream_mutable, position + 4)[0]
        struct.pack_into(
            "<I",
            canonical_stream_mutable,
            position + 4,
            sheet_offset + len(inserted_records),
        )
    stable_stream = (
        bytes(canonical_stream_mutable[:first_global_eof])
        + inserted_records
        + bytes(canonical_stream_mutable[first_global_eof:])
    )
    required_sector_count = math.ceil(len(stable_stream) / sector_size)
    extra_sector_count = max(0, required_sector_count - len(canonical_chain))

    output_xls.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(canonical_v10_xls, output_xls)
    file_bytes = bytearray(output_xls.read_bytes())
    existing_sector_count = len(canonical_fat)
    appended_sector_ids = list(
        range(existing_sector_count, existing_sector_count + extra_sector_count)
    )
    fat_sector_ids = _fat_sector_ids(file_bytes, sector_size)
    fat_entries_per_sector = sector_size // 4
    if appended_sector_ids and appended_sector_ids[-1] >= len(fat_sector_ids) * fat_entries_per_sector:
        raise RuntimeError("Stable v1.1 OLE output requires an additional FAT sector")
    file_bytes.extend(b"\x00" * extra_sector_count * sector_size)
    output_chain = canonical_chain + appended_sector_ids
    padded_stream = stable_stream + b"\x00" * (len(output_chain) * sector_size - len(stable_stream))
    for index, sector in enumerate(output_chain):
        start = (sector + 1) * sector_size
        file_bytes[start:start + sector_size] = padded_stream[
            index * sector_size:(index + 1) * sector_size
        ]

    def set_fat_entry(sector: int, value: int) -> None:
        fat_sector = fat_sector_ids[sector // fat_entries_per_sector]
        offset = (fat_sector + 1) * sector_size + (sector % fat_entries_per_sector) * 4
        struct.pack_into("<I", file_bytes, offset, value)

    if appended_sector_ids:
        set_fat_entry(canonical_chain[-1], appended_sector_ids[0])
        for left, right in zip(appended_sector_ids, appended_sector_ids[1:]):
            set_fat_entry(left, right)
        set_fat_entry(appended_sector_ids[-1], olefile.ENDOFCHAIN)

    directory_entry_index = canonical_entry.sid
    directory_sector = canonical_dir_chain[directory_entry_index // (sector_size // 128)]
    directory_offset = (
        (directory_sector + 1) * sector_size
        + (directory_entry_index % (sector_size // 128)) * 128
    )
    struct.pack_into("<Q", file_bytes, directory_offset + 120, len(stable_stream))
    output_xls.write_bytes(file_bytes)
    return output_xls


def _inventory_errors(
    label: str,
    inventory: dict[str, Any],
    expected_inventory: dict[str, Any],
    variant: str,
) -> list[str]:
    errors = [f"{label}: {error}" for error in inventory.get("errors", [])]
    errors.extend(
        f"{label}: {error}"
        for error in compare_named_range_inventories(
            expected_inventory,
            inventory,
            required_names(variant),
        )
    )
    return errors


def _validated_v11_inventory_pair(
    xls_path: Path,
    workbook,
    manifest: dict[str, Any],
    expected_inventory: dict[str, Any] | None = None,
) -> tuple[dict[str, Any], dict[str, Any]]:
    variant = "with_skill"
    raw_xls_inventory = validate_xls_named_range_inventory(
        xls_path,
        manifest["anchors"],
        variant,
    )
    xlsx_inventory = validate_named_range_inventory(
        workbook,
        manifest["anchors"],
        variant,
    )
    inventory_errors = [
        *raw_xls_inventory.get("errors", []),
        *xlsx_inventory.get("errors", []),
        *compare_xls_and_xlsx_named_ranges(
            raw_xls_inventory,
            xlsx_inventory,
            required_names(variant),
        ),
        *compare_named_range_inventories(
            raw_xls_inventory,
            xlsx_inventory,
            required_names(variant),
        ),
    ]
    if expected_inventory is not None:
        inventory_errors.extend(
            _inventory_errors(
                "controlled XLS",
                raw_xls_inventory,
                expected_inventory,
                variant,
            )
        )
        inventory_errors.extend(
            _inventory_errors(
                "controlled XLSX",
                xlsx_inventory,
                expected_inventory,
                variant,
            )
        )
    if inventory_errors:
        raise RuntimeError(
            "Controlled v1.1 named-range inventory is invalid: "
            + "; ".join(sorted(set(inventory_errors)))
        )
    return raw_xls_inventory, xlsx_inventory


def _roundtrip_xlsx_with_openpyxl(
    source_xlsx: Path,
    temp_dir: Path,
    soffice: str,
) -> tuple[Path, Path, Any]:
    """Reproduce the current-environment post-build XLS round trip."""
    normalized_xlsx = temp_dir / "openpyxl-xlsx" / source_xlsx.name
    normalized_xlsx.parent.mkdir(parents=True, exist_ok=True)
    load_workbook(source_xlsx, data_only=False).save(normalized_xlsx)
    controlled_xls = convert_to_xls(
        normalized_xlsx,
        temp_dir / "controlled-xls",
        soffice,
    )
    normalize_libreoffice_print_title_records(controlled_xls)
    normalize_xls_summary_information(controlled_xls)
    controlled_xlsx = convert_to_xlsx(
        controlled_xls,
        temp_dir / "controlled-xlsx",
        soffice,
    )
    return controlled_xls, controlled_xlsx, load_workbook(controlled_xlsx, data_only=False)


def build_controlled_v11_baseline(
    temp_dir: Path,
    soffice: str,
) -> ControlledV11Baseline:
    """Create a v1.1 workbook from canonical v1.0 in the current environment.

    The cache is scoped to the caller's temporary directory. This keeps all
    artifacts disposable while ensuring repeated comparisons in one process do
    not start another LibreOffice conversion pipeline.
    """
    temp_dir = Path(temp_dir).expanduser().resolve()
    cache_key = (str(temp_dir), str(soffice))
    cached = _BASELINE_CACHE.get(cache_key)
    if cached is not None:
        return cached

    temp_dir.mkdir(parents=True, exist_ok=True)
    base_manifest = load_manifest(V10_MANIFEST)
    v11_manifest = load_manifest(V11_MANIFEST)
    variant = "with_skill"
    expected_locations = expected_named_range_locations(base_manifest["structure"], variant)
    expected_inventory = {
        "locations": {
            name: location.to_dict() for name, location in expected_locations.items()
        }
    }

    source_xlsx = convert_to_xlsx(V10_TEMPLATE, temp_dir / "v10-source-xlsx", soffice)
    named_xlsx = temp_dir / "v11-controlled-source" / "template.xlsx"
    named_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(source_xlsx, data_only=False)
    for name in required_names(variant):
        set_named_range_from_location(workbook, expected_locations[name])
    workbook.save(named_xlsx)

    controlled_xls = convert_to_xls(named_xlsx, temp_dir / "v11-controlled-xls", soffice)
    normalize_libreoffice_print_title_records(controlled_xls)
    normalize_xls_summary_information(controlled_xls)
    controlled_xlsx = convert_to_xlsx(
        controlled_xls,
        temp_dir / "v11-controlled-roundtrip-xlsx",
        soffice,
    )
    controlled_workbook = load_workbook(controlled_xlsx, data_only=False)

    raw_xls_inventory, xlsx_inventory = _validated_v11_inventory_pair(
        controlled_xls,
        controlled_workbook,
        v11_manifest,
        expected_inventory,
    )

    baseline = ControlledV11Baseline(
        controlled_v11_xls=controlled_xls,
        controlled_v11_xlsx=controlled_xlsx,
        controlled_workbook=controlled_workbook,
        raw_xls_inventory=raw_xls_inventory,
        xlsx_inventory=xlsx_inventory,
        expected_locations=expected_locations,
    )
    _BASELINE_CACHE[cache_key] = baseline
    return baseline


def build_controlled_v11_candidate_roundtrip(
    template_xls: Path,
    temp_dir: Path,
    soffice: str,
) -> ControlledV11Roundtrip:
    """Round-trip a committed v1.1 candidate through the controlled path."""
    template_xls = Path(template_xls).expanduser().resolve()
    temp_dir = Path(temp_dir).expanduser().resolve()
    cache_key = (str(template_xls), str(temp_dir), str(soffice))
    cached = _CANDIDATE_CACHE.get(cache_key)
    if cached is not None:
        return cached

    temp_dir.mkdir(parents=True, exist_ok=True)
    v11_manifest = load_manifest(V11_MANIFEST)
    source_xlsx = convert_to_xlsx(template_xls, temp_dir / "source-xlsx", soffice)
    controlled_xls, controlled_xlsx, controlled_workbook = _roundtrip_xlsx_with_openpyxl(
        source_xlsx,
        temp_dir,
        soffice,
    )
    raw_xls_inventory, xlsx_inventory = _validated_v11_inventory_pair(
        controlled_xls,
        controlled_workbook,
        v11_manifest,
    )
    candidate = ControlledV11Roundtrip(
        controlled_xls=controlled_xls,
        controlled_xlsx=controlled_xlsx,
        controlled_workbook=controlled_workbook,
        raw_xls_inventory=raw_xls_inventory,
        xlsx_inventory=xlsx_inventory,
    )
    _CANDIDATE_CACHE[cache_key] = candidate
    return candidate
