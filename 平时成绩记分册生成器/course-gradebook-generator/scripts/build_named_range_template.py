#!/usr/bin/env python3
"""Build the v1.1 .xls package from the protected v1.0 canonical template."""

from __future__ import annotations

import argparse
import hashlib
import os
import re
import shutil
import subprocess
import tempfile
import uuid
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

from named_range_contracts import required_names
from named_range_utils import (
    compare_named_range_inventories,
    validate_named_range_inventory,
)
from package_common import (
    V10_MANIFEST,
    V10_TEMPLATE,
    V11_PACKAGE_DIR,
    sha256_file,
    validate_canonical_baselines,
)
from named_range_template_baseline import (
    build_controlled_v11_baseline,
    build_controlled_v11_candidate_roundtrip,
)
from validate_template import controlled_roundtrip_paths, convert_to_format, find_soffice
from xls_named_range_utils import (
    compare_xls_and_xlsx_named_ranges,
    validate_xls_named_range_inventory,
)


def replace_directory_atomically(stage: Path, output_dir: Path) -> None:
    """Exchange a fully validated package directory with rollback on failure."""
    stage = stage.expanduser().resolve()
    output_dir = output_dir.expanduser().resolve()
    if not stage.is_dir():
        raise RuntimeError(f"Directory swap stage is not a directory: {stage}")
    if stage.parent != output_dir.parent:
        raise RuntimeError("Directory swap stage and output must share the same parent directory")
    if output_dir.exists() and not output_dir.is_dir():
        raise RuntimeError(f"Directory swap output is not a directory: {output_dir}")

    backup = output_dir.parent / f".{output_dir.name}.{uuid.uuid4().hex}.backup"
    moved_old = False
    try:
        if output_dir.exists():
            os.replace(str(output_dir), str(backup))
            moved_old = True
        if os.environ.get("GRADEBOOK_TEST_FAIL_DIRECTORY_SWAP") == "1":
            raise OSError("Injected directory swap failure")
        os.replace(str(stage), str(output_dir))
    except Exception as exc:
        restore_error = None
        if moved_old:
            try:
                if output_dir.exists():
                    if output_dir.is_dir():
                        shutil.rmtree(output_dir)
                    else:
                        output_dir.unlink()
                os.replace(str(backup), str(output_dir))
            except Exception as restore_exc:
                restore_error = restore_exc
        if restore_error is not None:
            raise RuntimeError(
                f"Directory swap failed: {exc}; restoring previous package failed: {restore_error}"
            ) from exc
        raise RuntimeError(f"Directory swap failed: {exc}") from exc
    else:
        if moved_old:
            try:
                shutil.rmtree(backup)
            except Exception as cleanup_exc:
                raise RuntimeError(f"Directory swap succeeded but backup cleanup failed: {cleanup_exc}") from cleanup_exc
    finally:
        if stage.exists():
            shutil.rmtree(stage)


def _workbook_layout_signature(workbook) -> dict[str, Any]:
    def cell_signature(cell) -> tuple[Any, ...]:
        return (cell.coordinate, cell.value, cell.data_type, cell.style_id, cell.number_format)

    def dimension_signature(value) -> float | None:
        return None if value is None else round(float(value), 0)

    return {
        "sheetnames": list(workbook.sheetnames),
        "states": {sheet.title: sheet.sheet_state for sheet in workbook.worksheets},
        "sheets": {
            sheet.title: {
                "cells": [
                    cell_signature(cell)
                    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)
                    for cell in row
                    if cell.value is not None or cell.has_style
                ],
                "merged": sorted(str(item).upper() for item in sheet.merged_cells.ranges),
                "column_dimensions": sorted(
                    (str(key), dimension_signature(value.width), bool(value.hidden), int(value.outlineLevel), bool(value.collapsed))
                    for key, value in sheet.column_dimensions.items()
                ),
                "row_dimensions": sorted(
                    (str(key), dimension_signature(value.height), bool(value.hidden), int(value.outlineLevel), bool(value.collapsed))
                    for key, value in sheet.row_dimensions.items()
                    if value.height is not None or value.hidden or value.outlineLevel or value.collapsed
                ),
                "page_setup": (
                    sheet.page_setup.orientation,
                    sheet.page_setup.paperSize,
                    sheet.page_setup.scale,
                    sheet.page_margins.left,
                    sheet.page_margins.right,
                    sheet.page_margins.top,
                    sheet.page_margins.bottom,
                    str(sheet.print_area or ""),
                    str(sheet.freeze_panes or ""),
                ),
                "protection": (
                    bool(sheet.protection.sheet),
                    bool(sheet.protection.formatCells),
                    bool(sheet.protection.formatColumns),
                    bool(sheet.protection.formatRows),
                    bool(sheet.protection.insertColumns),
                    bool(sheet.protection.insertRows),
                    bool(sheet.protection.deleteColumns),
                    bool(sheet.protection.deleteRows),
                ),
            }
            for sheet in workbook.worksheets
        },
        "workbook_protection": (
            bool(getattr(workbook.security, "lockStructure", False)),
            bool(getattr(workbook.security, "lockWindows", False)),
            bool(getattr(workbook.security, "lockRevision", False)),
            str(getattr(workbook.security, "workbookPassword", "") or ""),
            str(getattr(workbook.security, "revisionsPassword", "") or ""),
        ),
    }


def _pdf_signature(path: Path) -> tuple[int, int]:
    payload = path.read_bytes()
    if not payload:
        raise RuntimeError(f"LibreOffice created an empty PDF: {path}")
    return len(re.findall(rb"/Type\s*/Page(?:\s|/|>)", payload)), len(payload)


def _convert(soffice: str, source: Path, output_dir: Path, target: str) -> Path:
    return convert_to_format(source, output_dir, target, soffice)


def _assert_fixed_v10_baseline(
    source_xlsx: Path,
    canonical_xlsx: Path,
    soffice: str,
    source: Path,
    temp: Path,
) -> None:
    if _workbook_layout_signature(load_workbook(source_xlsx, data_only=False)) != _workbook_layout_signature(
        load_workbook(canonical_xlsx, data_only=False)
    ):
        raise RuntimeError(
            "Named-range build rejected a source whose protected layout or formatting differs from canonical v1.0"
        )
    source_pdf = _convert(soffice, source, temp / "source-pdf", "pdf")
    canonical_pdf = _convert(soffice, V10_TEMPLATE, temp / "canonical-pdf", "pdf")
    if _pdf_signature(source_pdf)[0] != _pdf_signature(canonical_pdf)[0]:
        raise RuntimeError("Named-range build rejected a source whose PDF page layout differs from canonical v1.0")


def _write_manifest(
    source_manifest: Path,
    target_manifest: Path,
    template_sha: str,
    *,
    baseline_prefix: str | None = None,
) -> None:
    manifest = yaml.safe_load(source_manifest.read_text(encoding="utf-8")) or {}
    manifest["fingerprint"]["sha256"] = template_sha
    manifest["fingerprint"]["value"] = template_sha
    if baseline_prefix:
        manifest["template"]["base_manifest"] = f"{baseline_prefix}/manifest.yaml"
        manifest["template"]["base_template"] = f"{baseline_prefix}/template.xls"
    target_manifest.write_text(
        yaml.safe_dump(manifest, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )


def _inventory_has_managed_names(inventory: dict[str, Any]) -> bool:
    return bool(
        inventory.get("locations")
        or inventory.get("duplicate")
        or inventory.get("invalid_names")
        or inventory.get("unexpected")
        or inventory.get("scope_errors")
        or inventory.get("broken")
        or inventory.get("destination_errors")
        or inventory.get("shape_errors")
        or inventory.get("relationship_errors")
    )


def _validate_complete_named_inventory(
    label: str,
    inventory: dict[str, Any],
    expected_inventory: dict[str, Any],
) -> list[str]:
    errors = list(inventory.get("errors", []))
    errors.extend(
        compare_named_range_inventories(
            expected_inventory,
            inventory,
            required_names("with_skill"),
        )
    )
    if errors:
        return [f"{label}: {error}" for error in sorted(set(errors))]
    return []


def build_template(source: Path, output_dir: Path, force: bool = False) -> dict[str, Any]:
    source = source.expanduser().resolve()
    output_dir = output_dir.expanduser().resolve()
    if not source.exists():
        raise RuntimeError(f"Source template not found: {source}")
    if output_dir.exists() and not force:
        raise RuntimeError(f"Refusing to overwrite existing output package: {output_dir}; use --force explicitly")
    validate_canonical_baselines(require_v11=True)
    soffice = find_soffice()
    source_manifest = source.parent / "manifest.yaml"
    if not source_manifest.exists():
        source_manifest = source.parent.parent / "v1.1.0" / "manifest.yaml"
    with tempfile.TemporaryDirectory(prefix="gradebook-named-template-") as temp_name:
        temp = Path(temp_name)
        source_manifest_v11 = V11_PACKAGE_DIR / "manifest.yaml"
        manifest_data = yaml.safe_load(source_manifest_v11.read_text(encoding="utf-8")) or {}
        source_xls_inventory = validate_xls_named_range_inventory(
            source,
            manifest_data["anchors"],
            "with_skill",
        )
        source_xlsx = _convert(soffice, source, temp / "source-xlsx", "xlsx")
        canonical_xlsx = _convert(soffice, V10_TEMPLATE, temp / "canonical-xlsx", "xlsx")
        _assert_fixed_v10_baseline(source_xlsx, canonical_xlsx, soffice, source, temp)
        controlled_baseline = build_controlled_v11_baseline(
            temp / "controlled-v11-baseline",
            soffice,
        )
        expected_locations = controlled_baseline.expected_locations
        expected_inventory = {
            "locations": {
                name: location.to_dict() for name, location in expected_locations.items()
            }
        }
        source_xlsx_inventory = validate_named_range_inventory(
            load_workbook(source_xlsx, data_only=False),
            manifest_data["anchors"],
            "with_skill",
        )
        source_has_names = _inventory_has_managed_names(source_xls_inventory)
        source_xlsx_has_names = _inventory_has_managed_names(source_xlsx_inventory)
        if not source_has_names and not source_xlsx_has_names:
            source_named_state = "v1.0 seed"
        elif not source_has_names or not source_xlsx_has_names:
            raise RuntimeError(
                "Named-range build rejected a partial managed-name inventory: "
                "raw XLS and round-trip XLSX do not contain the same managed-name state"
            )
        else:
            expected_count = len(required_names("with_skill"))
            if (
                source_xls_inventory.get("actual_count") != expected_count
                or source_xlsx_inventory.get("actual_count") != expected_count
            ):
                raise RuntimeError(
                    "Named-range build rejected a partial managed-name inventory: "
                    f"expected {expected_count} complete names, got "
                    f"raw={source_xls_inventory.get('actual_count')}, "
                    f"roundtrip={source_xlsx_inventory.get('actual_count')}"
                )
            source_errors = _validate_complete_named_inventory(
                "source XLS",
                source_xls_inventory,
                expected_inventory,
            )
            source_errors.extend(
                _validate_complete_named_inventory(
                    "source round-trip XLSX",
                    source_xlsx_inventory,
                    expected_inventory,
                )
            )
            source_errors.extend(
                compare_xls_and_xlsx_named_ranges(
                    source_xls_inventory,
                    source_xlsx_inventory,
                    required_names("with_skill"),
                )
            )
            source_errors.extend(
                compare_named_range_inventories(
                    source_xls_inventory,
                    source_xlsx_inventory,
                    required_names("with_skill"),
                )
            )
            if source_errors:
                raise RuntimeError(
                    "Named-range build rejected an invalid managed-name inventory: "
                    + "; ".join(sorted(set(source_errors)))
                )
            source_named_state = "v1.1 canonical"
        if source_named_state == "v1.0 seed":
            named_xls = controlled_baseline.controlled_v11_xls
            named_roundtrip_xlsx = controlled_baseline.controlled_v11_xlsx
        else:
            # A complete v1.1 source is already a canonical semantic package.
            # Validate its raw and round-trip inventories, then preserve its
            # bytes instead of silently rewriting names through LibreOffice.
            named_xls = source
            named_roundtrip_xlsx = build_controlled_v11_candidate_roundtrip(
                named_xls,
                temp / "controlled-v11-candidate",
                soffice,
            ).controlled_xlsx

        package_manifest = output_dir / "manifest.yaml"
        source_v11_manifest = V11_PACKAGE_DIR / "manifest.yaml"
        manifest_source = source_v11_manifest if source_v11_manifest.exists() else source_manifest
        manifest_data = yaml.safe_load(manifest_source.read_text(encoding="utf-8")) or {}
        xlsx_inventory = validate_named_range_inventory(
            load_workbook(named_roundtrip_xlsx, data_only=False),
            manifest_data["anchors"],
            "with_skill",
        )
        xls_inventory = validate_xls_named_range_inventory(named_xls, manifest_data["anchors"], "with_skill")
        if xlsx_inventory["errors"] or xls_inventory["errors"]:
            raise RuntimeError("Named range build failed: " + "; ".join(xlsx_inventory["errors"] + xls_inventory["errors"]))
        expected_differences = compare_named_range_inventories(
            expected_inventory,
            xls_inventory,
            required_names("with_skill"),
        )
        expected_differences.extend(
            compare_named_range_inventories(
                expected_inventory,
                xlsx_inventory,
                required_names("with_skill"),
            )
        )
        if expected_differences:
            raise RuntimeError(
                "Named range build changed the canonical managed destinations: "
                + "; ".join(sorted(set(expected_differences)))
            )
        differences = compare_xls_and_xlsx_named_ranges(xls_inventory, xlsx_inventory, required_names("with_skill"))
        differences.extend(compare_named_range_inventories(xls_inventory, xlsx_inventory, required_names("with_skill")))
        if differences:
            raise RuntimeError("Raw XLS and round-trip XLSX named-range inventories differ: " + "; ".join(sorted(set(differences))))
        _, controlled_v10_xlsx = controlled_roundtrip_paths(
            V10_TEMPLATE,
            temp / "controlled-v10-baseline",
            soffice,
        )
        if _workbook_layout_signature(load_workbook(controlled_v10_xlsx, data_only=False)) != _workbook_layout_signature(
            load_workbook(controlled_baseline.controlled_v11_xlsx, data_only=False)
        ):
            raise RuntimeError("Named-range build changed protected workbook layout or formatting")
        if _workbook_layout_signature(load_workbook(controlled_baseline.comparison_v11_xlsx, data_only=False)) != _workbook_layout_signature(
            load_workbook(named_roundtrip_xlsx, data_only=False)
        ):
            raise RuntimeError("Named-range template differs from the controlled v1.1 baseline")

        baseline_pdf = _convert(soffice, V10_TEMPLATE, temp / "baseline-pdf", "pdf")
        named_pdf = _convert(soffice, named_xls, temp / "named-pdf", "pdf")
        if _pdf_signature(baseline_pdf)[0] != _pdf_signature(named_pdf)[0]:
            raise RuntimeError(
                f"Named-range build changed PDF rendering: baseline={_pdf_signature(baseline_pdf)}, named={_pdf_signature(named_pdf)}"
            )

        stage = temp / "package"
        stage.mkdir()
        staged_template = stage / "template.xls"
        shutil.copy2(named_xls, staged_template)
        template_sha = sha256_file(staged_template)
        custom_output_package = output_dir.resolve() != V11_PACKAGE_DIR.resolve()
        baseline_prefix = "baseline-v1.0.0" if custom_output_package else None
        _write_manifest(
            manifest_source,
            stage / "manifest.yaml",
            template_sha,
            baseline_prefix=baseline_prefix,
        )
        if custom_output_package:
            baseline_stage = stage / baseline_prefix
            baseline_stage.mkdir(parents=True, exist_ok=True)
            shutil.copy2(V10_MANIFEST, baseline_stage / "manifest.yaml")
            shutil.copy2(V10_TEMPLATE, baseline_stage / "template.xls")
        changelog = manifest_source.parent / "CHANGELOG.md"
        if changelog.exists():
            shutil.copy2(changelog, stage / "CHANGELOG.md")
        else:
            (stage / "CHANGELOG.md").write_text("# 1.1.0\n\n- Add workbook-level managed named ranges.\n", encoding="utf-8")
        output_dir.parent.mkdir(parents=True, exist_ok=True)
        exchange_stage = output_dir.parent / f".{output_dir.name}.{uuid.uuid4().hex}.stage"
        try:
            shutil.copytree(stage, exchange_stage)
            replace_directory_atomically(exchange_stage, output_dir)
        finally:
            if exchange_stage.exists():
                shutil.rmtree(exchange_stage)
        return {
            "output": str(output_dir),
            "template_sha256": template_sha,
            "v10_template_sha256": sha256_file(V10_TEMPLATE),
            "named_range_count": len(xls_inventory["locations"]),
            "pdf_signature": _pdf_signature(named_pdf),
        }


def main() -> int:
    parser = argparse.ArgumentParser(description="Build the course-gradebook v1.1 named-range template package.")
    parser.add_argument("--source", default=str(V10_TEMPLATE))
    parser.add_argument("--output-dir", default=str(V11_PACKAGE_DIR))
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()
    result = build_template(Path(args.source), Path(args.output_dir), args.force)
    print(yaml.safe_dump(result, allow_unicode=True, sort_keys=False), end="")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
