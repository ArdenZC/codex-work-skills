from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import subprocess
import sys
import tempfile
import unicodedata
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from xlrd import open_workbook

from named_range_contracts import required_names
from named_range_utils import (
    compare_named_range_inventories,
    expected_named_range_locations,
    validate_named_range_inventory,
)
from package_common import (
    DEFAULT_MANIFEST,
    V10_MANIFEST,
    V10_TEMPLATE,
    V11_TEMPLATE,
    anchor_mode,
    canonical_template_for_mode,
    column_number,
    ensure_supported_major,
    load_manifest,
    manifest_template_path,
    resolve_template_package,
    validate_canonical_baselines,
    validate_template_package_identity,
    validate_manifest_contract,
)
from xls_named_range_utils import compare_xls_and_xlsx_named_ranges, validate_xls_named_range_inventory


class TemplateValidationError(RuntimeError):
    def __init__(self, report: dict[str, Any]):
        self.report = report
        super().__init__("; ".join(report.get("errors", [])) or "Template validation failed")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def find_soffice() -> str:
    candidates = [
        shutil.which("soffice"),
        shutil.which("libreoffice"),
        shutil.which("soffice.com"),
        r"C:\Program Files\LibreOffice\program\soffice.com",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/usr/bin/libreoffice",
        "/usr/local/bin/libreoffice",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(candidate)
    raise RuntimeError("LibreOffice/soffice was not found; install LibreOffice or use Windows Excel COM for generation.")


def convert_to_format(source: Path, out_dir: Path, target_format: str, soffice: str) -> Path:
    target_format = str(target_format).lower().lstrip(".")
    out_dir.mkdir(parents=True, exist_ok=True)
    expected = out_dir / f"{source.stem}.{target_format}"
    existing_candidates = sorted(
        path for path in out_dir.iterdir() if path.is_file() and path.stem == source.stem
    )
    if existing_candidates:
        raise RuntimeError(
            f"LibreOffice conversion output directory is not fresh for source={source} target={target_format}: "
            f"{[str(path) for path in existing_candidates]}"
        )
    proc = subprocess.run(
        [soffice, "--headless", "--convert-to", target_format, "--outdir", str(out_dir), str(source)],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=120,
    )
    if proc.returncode != 0:
        raise RuntimeError(
            f"LibreOffice conversion failed for source={source} expected={expected}\n"
            f"stdout:\n{proc.stdout}\nstderr:\n{proc.stderr}"
        )
    matches = sorted(path for path in out_dir.iterdir() if path.is_file() and path.stem == source.stem)
    if len(matches) != 1 or matches[0] != expected or not expected.is_file():
        raise RuntimeError(
            f"LibreOffice conversion did not produce exactly the expected {target_format.upper()} output "
            f"for source={source} expected={expected} candidates={[str(path) for path in matches]}\n"
            f"stdout:\n{proc.stdout}\nstderr:\n{proc.stderr}"
        )
    if expected.stat().st_size <= 0:
        raise RuntimeError(
            f"LibreOffice produced an empty output for source={source} expected={expected}\n"
            f"stdout:\n{proc.stdout}\nstderr:\n{proc.stderr}"
        )
    return expected


def convert_to_xlsx(source: Path, out_dir: Path, soffice: str) -> Path:
    return convert_to_format(source, out_dir, "xlsx", soffice)


def convert_to_xls(source: Path, out_dir: Path, soffice: str) -> Path:
    return convert_to_format(source, out_dir, "xls", soffice)


def controlled_roundtrip_paths(source_xls: Path, out_dir: Path, soffice: str) -> tuple[Path, Path]:
    """Create the platform-specific template-to-output font baseline."""
    source_xlsx = convert_to_xlsx(source_xls, out_dir / "source-xlsx", soffice)
    # The generator loads the converted template with openpyxl and saves it
    # before LibreOffice writes the final XLS. Reproduce that style pipeline so
    # fallback signatures are measured against the actual supported conversion.
    normalized_xlsx = out_dir / "openpyxl-xlsx" / source_xlsx.name
    normalized_xlsx.parent.mkdir(parents=True, exist_ok=True)
    load_workbook(source_xlsx, data_only=False).save(normalized_xlsx)
    roundtrip_xls = convert_to_xls(normalized_xlsx, out_dir / "roundtrip-xls", soffice)
    return roundtrip_xls, convert_to_xlsx(roundtrip_xls, out_dir / "roundtrip-xlsx", soffice)


def controlled_roundtrip_xlsx(source_xls: Path, out_dir: Path, soffice: str) -> Path:
    return controlled_roundtrip_paths(source_xls, out_dir, soffice)[1]


def controlled_font_reference_roundtrip_xlsx(
    source_xlsx: Path,
    reference_xlsx: Path,
    replacements: list[tuple[str, str, str]],
    out_dir: Path,
    soffice: str,
) -> Path:
    """Round-trip content with canonical fonts restored at controlled coordinates."""
    source_workbook = load_workbook(source_xlsx, data_only=False)
    reference_workbook = load_workbook(reference_xlsx, data_only=False)
    for sheet_name, target_address, reference_address in replacements:
        if sheet_name not in source_workbook.sheetnames or sheet_name not in reference_workbook.sheetnames:
            continue
        target_ws = source_workbook[sheet_name]
        reference_ws = reference_workbook[sheet_name]
        if target_address in target_ws and reference_address in reference_ws:
            target_ws[target_address].font = copy(reference_ws[reference_address].font)
    normalized_xlsx = out_dir / "font-reference-xlsx" / source_xlsx.name
    normalized_xlsx.parent.mkdir(parents=True, exist_ok=True)
    source_workbook.save(normalized_xlsx)
    roundtrip_xls = convert_to_xls(normalized_xlsx, out_dir / "roundtrip-xls", soffice)
    return convert_to_xlsx(roundtrip_xls, out_dir / "roundtrip-xlsx", soffice)


def controlled_content_roundtrip_paths(
    reference_xlsx: Path,
    content_xlsx: Path,
    sheet_name: str,
    out_dir: Path,
    soffice: str,
) -> tuple[Path, Path]:
    """Round-trip canonical styles with the candidate workbook's cell contents."""
    reference_workbook = load_workbook(reference_xlsx, data_only=False)
    content_workbook = load_workbook(content_xlsx, data_only=False)
    if sheet_name in reference_workbook.sheetnames and sheet_name in content_workbook.sheetnames:
        reference_ws = reference_workbook[sheet_name]
        content_ws = content_workbook[sheet_name]
        for row in range(1, content_ws.max_row + 1):
            for column in range(1, content_ws.max_column + 1):
                target = reference_ws.cell(row, column)
                if isinstance(target, MergedCell):
                    continue
                target.value = content_ws.cell(row, column).value
    normalized_xlsx = out_dir / "content-xlsx" / reference_xlsx.name
    normalized_xlsx.parent.mkdir(parents=True, exist_ok=True)
    reference_workbook.save(normalized_xlsx)
    roundtrip_xls = convert_to_xls(normalized_xlsx, out_dir / "roundtrip-xls", soffice)
    return roundtrip_xls, convert_to_xlsx(roundtrip_xls, out_dir / "roundtrip-xlsx", soffice)


def xls_font_identity(path: Path, sheet_name: str, address: str) -> tuple[Any, ...] | None:
    """Read the source XLS font identity before LibreOffice conversion can normalize it."""
    workbook = open_workbook(str(path), formatting_info=True)
    sheet = workbook.sheet_by_name(sheet_name)
    match = re.fullmatch(r"([A-Z]+)(\d+)", str(address).upper())
    if match is None:
        raise ValueError(f"Invalid XLS cell address: {address}")
    column = column_number(match.group(1)) - 1
    row = int(match.group(2)) - 1
    if row < 0 or row >= sheet.nrows or column < 0 or column >= sheet.ncols:
        return None
    try:
        xf = workbook.xf_list[sheet.cell_xf_index(row, column)]
    except (IndexError, ValueError):
        return None
    font = workbook.font_list[xf.font_index]
    return _font_name_signature(font.name), font.character_set, font.family


_FONT_NAME_ALIASES = {
    "simsun": "simsun",
    "mssimsun": "simsun",
    "宋体": "simsun",
    "nsimsun": "nsimsun",
    "新宋体": "nsimsun",
    "simhei": "simhei",
    "黑体": "simhei",
    "microsoftyahei": "microsoftyahei",
    "microsoftyaheiui": "microsoftyahei",
    "微软雅黑": "microsoftyahei",
    "kaiti": "kaiti",
    "楷体": "kaiti",
    "fangsong": "fangsong",
    "仿宋": "fangsong",
    "dengxian": "dengxian",
    "等线": "dengxian",
}

def _font_name_signature(value: Any) -> str:
    normalized = unicodedata.normalize("NFKC", str(value or "")).strip().casefold()
    compact = re.sub(r"[\s_-]+", "", normalized)
    return _FONT_NAME_ALIASES.get(compact, compact)


def _font_metadata_signature(font) -> tuple[Any, ...]:
    return font.charset, font.family, font.scheme


def _font_signatures_match(left: tuple[Any, ...], right: tuple[Any, ...]) -> bool:
    """Compare the complete font identity, including the declared family name."""
    return len(left) == len(right) and left == right


def _cell_format_signatures_match(left: dict[str, Any], right: dict[str, Any]) -> bool:
    if set(left) != set(right):
        return False
    for key in left:
        if key == "font":
            if not _font_signatures_match(left[key], right[key]):
                return False
        elif left[key] != right[key]:
            return False
    return True


def _cell_format_signature(cell) -> dict[str, Any]:
    def color_signature(color) -> tuple[Any, ...]:
        if color is None:
            return ()
        return (
            color.type,
            color.rgb,
            color.indexed,
            color.auto,
            color.theme,
            color.tint,
        )

    def side_signature(side) -> tuple[Any, ...]:
        if side is None:
            return ()
        return side.style, color_signature(side.color)

    font = cell.font
    fill = cell.fill
    border = cell.border
    alignment = cell.alignment
    protection = cell.protection
    font_name = _font_name_signature(font.name)
    font_charset, font_family, font_scheme = _font_metadata_signature(font)
    return {
        "number_format": cell.number_format,
        # Keep font identity while normalizing common locale-specific aliases from XLS round trips.
        "font": (
            font_name,
            font_charset,
            font_family,
            font_scheme,
            font.sz,
            bool(font.b),
            bool(font.i),
            font.u,
            bool(font.strike),
            bool(font.outline),
            bool(font.shadow),
            font.vertAlign,
            color_signature(font.color),
        ),
        "fill": (
            fill.patternType,
            color_signature(fill.fgColor),
            color_signature(fill.bgColor),
        ),
        "border": (
            border.outline,
            border.diagonalUp,
            border.diagonalDown,
            side_signature(border.left),
            side_signature(border.right),
            side_signature(border.top),
            side_signature(border.bottom),
            side_signature(border.diagonal),
            side_signature(border.start),
            side_signature(border.end),
        ),
        "alignment": (
            alignment.horizontal,
            alignment.vertical,
            alignment.textRotation,
            bool(alignment.wrapText),
            bool(alignment.shrinkToFit),
            alignment.indent or 0,
            alignment.relativeIndent or 0,
            alignment.justifyLastLine,
            alignment.readingOrder or 0,
        ),
        "protection": (bool(protection.locked), bool(protection.hidden)),
    }


def _dimension_signature(value: Any) -> float | None:
    if value is None:
        return None
    # LibreOffice can shift XLS column widths by hundredths during a round trip.
    return round(float(value), 1)


def _page_setup_value(value: Any) -> Any:
    if isinstance(value, float):
        # LibreOffice can introduce a few-millionths drift in inch-based margins.
        return round(value, 4)
    return value


def _protection_signature(protection) -> dict[str, Any]:
    signature: dict[str, Any] = {}
    for name in getattr(protection, "__attrs__", ()):
        value = getattr(protection, name, None)
        if any(token in name.lower() for token in ("password", "hash", "salt")):
            present = value not in (None, "")
            signature[name] = {
                "present": present,
                "digest": hashlib.sha256(str(value).encode("utf-8")).hexdigest() if present else "",
            }
        else:
            signature[name] = value
    return signature


def _header_footer_item_signature(item) -> dict[str, dict[str, Any]]:
    return {
        side: {
            name: getattr(getattr(item, side, None), name, None)
            for name in ("text", "size", "font", "color")
        }
        for side in ("left", "center", "right")
    }


def _header_footer_signature(sheet) -> dict[str, Any]:
    header_footer = sheet.HeaderFooter
    return {
        "flags": {
            name: getattr(header_footer, name, None)
            for name in ("differentOddEven", "differentFirst", "scaleWithDoc", "alignWithMargins")
        },
        "odd_header": _header_footer_item_signature(header_footer.oddHeader),
        "odd_footer": _header_footer_item_signature(header_footer.oddFooter),
        "even_header": _header_footer_item_signature(header_footer.evenHeader),
        "even_footer": _header_footer_item_signature(header_footer.evenFooter),
        "first_header": _header_footer_item_signature(header_footer.firstHeader),
        "first_footer": _header_footer_item_signature(header_footer.firstFooter),
    }


def _page_setup_signature(sheet) -> dict[str, Any]:
    page_setup = sheet.page_setup
    margins = sheet.page_margins
    print_options = sheet.print_options
    page_setup_properties = sheet.sheet_properties.pageSetUpPr
    return {
        "page_setup": {
            name: _page_setup_value(getattr(page_setup, name, None))
            for name in (
                "orientation",
                "paperSize",
                "scale",
                "fitToHeight",
                "fitToWidth",
                "firstPageNumber",
                "useFirstPageNumber",
                "paperHeight",
                "paperWidth",
                "pageOrder",
                "usePrinterDefaults",
                "blackAndWhite",
                "draft",
                "cellComments",
                "errors",
                "horizontalDpi",
                "verticalDpi",
                "copies",
                "autoPageBreaks",
                "fitToPage",
            )
        },
        "margins": {
            name: _page_setup_value(getattr(margins, name, None))
            for name in ("left", "right", "top", "bottom", "header", "footer")
        },
        "print_options": {
            name: getattr(print_options, name, None)
            for name in ("horizontalCentered", "verticalCentered", "headings", "gridLines", "gridLinesSet")
        },
        "page_setup_properties": {
            "fitToPage": getattr(page_setup_properties, "fitToPage", None) if page_setup_properties else None,
            "autoPageBreaks": getattr(page_setup_properties, "autoPageBreaks", None) if page_setup_properties else None,
        },
        "print_title_rows": str(sheet.print_title_rows or ""),
        "print_title_cols": str(sheet.print_title_cols or ""),
        "header_footer": _header_footer_signature(sheet),
    }


def _non_target_dimension_signature(value: Any) -> float | None:
    if value is None:
        return None
    # LibreOffice can shift compatibility-sheet dimensions by fractional units during an XLS round trip.
    return round(float(value))


def _non_target_sheet_signature(sheet) -> dict[str, Any]:
    max_row = sheet.max_row
    max_column = sheet.max_column
    cells = []
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
        cells.append(
            [
                {
                    "value": cell.value,
                    "data_type": cell.data_type,
                    "format": _cell_format_signature(cell),
                }
                for cell in row
            ]
        )
    return {
        "dimension": [max_row, max_column],
        "cells": cells,
        "merged": sorted(str(item).upper() for item in sheet.merged_cells.ranges),
        "column_widths": {
            key: {
                "width": _non_target_dimension_signature(value.width),
                "hidden": bool(value.hidden),
                "outline_level": value.outlineLevel,
                "collapsed": bool(value.collapsed),
            }
            for key, value in sheet.column_dimensions.items()
        },
        "row_heights": {
            str(key): {
                "height": _non_target_dimension_signature(value.height),
                "hidden": bool(value.hidden),
                "outline_level": value.outlineLevel,
                "collapsed": bool(value.collapsed),
            }
            for key, value in sheet.row_dimensions.items()
            if value.height is not None or value.hidden or value.outlineLevel or value.collapsed
        },
        "orientation": sheet.page_setup.orientation,
        "print_area": str(sheet.print_area or ""),
        "freeze_panes": str(sheet.freeze_panes or ""),
        "page_setup": _page_setup_signature(sheet),
        "protection": _protection_signature(sheet.protection),
    }


def _target_cell_format_signature(sheet) -> dict[str, dict[str, Any]]:
    return {
        cell.coordinate: _cell_format_signature(cell)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)
        for cell in row
    }


def _protected_target_values(sheet, manifest: dict[str, Any]) -> dict[str, Any]:
    structure = manifest["structure"]
    writable_cells = {
        str(cell).upper()
        for cell in [*structure.get("metadata", {}).values(), *structure.get("headers", {}).values()]
        if cell
    }
    data_start_row = int(structure["data_start_row"])
    total_column = column_number(structure["columns"]["total_score"])
    return {
        f"{get_column_letter(column)}{row}": sheet.cell(row, column).value
        for row in range(1, data_start_row)
        for column in range(1, total_column + 1)
        if f"{get_column_letter(column)}{row}" not in writable_cells
    }


def _signature_differences(
    left: Any,
    right: Any,
    path: str = "",
    font_matcher=None,
) -> list[dict[str, Any]]:
    differences: list[dict[str, Any]] = []
    if path.endswith(".font") and isinstance(left, tuple) and isinstance(right, tuple):
        if not _font_signatures_match(left, right) and not (
            font_matcher is not None and font_matcher(left, right, path)
        ):
            differences.append({"path": path, "expected": left, "actual": right})
    elif isinstance(left, dict) and isinstance(right, dict):
        for key in sorted(set(left) | set(right), key=str):
            child_path = f"{path}.{key}" if path else str(key)
            if key not in left:
                differences.append({"path": child_path, "expected": "<missing>", "actual": right[key]})
            elif key not in right:
                differences.append({"path": child_path, "expected": left[key], "actual": "<missing>"})
            else:
                differences.extend(_signature_differences(left[key], right[key], child_path, font_matcher))
    elif isinstance(left, (list, tuple)) and isinstance(right, (list, tuple)):
        if len(left) != len(right):
            differences.append({"path": path, "expected": left, "actual": right})
        else:
            for index, (expected, actual) in enumerate(zip(left, right)):
                differences.extend(
                    _signature_differences(expected, actual, f"{path}[{index}]", font_matcher)
                )
    elif left != right:
        differences.append({"path": path, "expected": left, "actual": right})
    return differences


def _font_signature_at_path(signature: dict[str, Any], path: str) -> tuple[Any, ...] | None:
    match = re.fullmatch(
        r"(target_cell_formats|writable_cell_formats)\.([A-Z]+\d+)\.font",
        path,
        flags=re.IGNORECASE,
    )
    if match is None:
        return None
    cell_formats = signature.get(match.group(1))
    if not isinstance(cell_formats, dict):
        return None
    cell_signature = cell_formats.get(match.group(2).upper())
    if not isinstance(cell_signature, dict):
        return None
    font_signature = cell_signature.get("font")
    return font_signature if isinstance(font_signature, tuple) else None


def controlled_font_fallback_matches(
    expected: tuple[Any, ...],
    actual: tuple[Any, ...],
    controlled: tuple[Any, ...] | None,
) -> bool:
    """Match only a complete font signature observed at the same controlled coordinate."""
    if controlled is None or controlled == expected:
        return False
    return actual != expected and actual == controlled


def _controlled_font_fallback_at_path(
    expected: tuple[Any, ...],
    actual: tuple[Any, ...],
    path: str,
    controlled_signature: dict[str, Any] | None,
) -> bool:
    """Allow a font difference only when the same coordinate differs in the controlled baseline."""
    if controlled_signature is None:
        return False
    controlled = _font_signature_at_path(controlled_signature, path)
    return controlled_font_fallback_matches(expected, actual, controlled)


def _workbook_signature(workbook, manifest: dict[str, Any]) -> dict[str, Any]:
    structure = manifest["structure"]
    sheet_name = structure["worksheet"]
    ws = workbook[sheet_name]
    label_cells = structure.get("header_label_cells", {})
    writable_cells = {
        str(cell)
        for cell in [
            *structure.get("metadata", {}).values(),
            *structure.get("headers", {}).values(),
        ]
        if cell
    }
    fixed_cells = [structure.get("title_cell"), *label_cells.values()]
    fixed_cells = [cell for cell in fixed_cells if cell and str(cell) not in writable_cells]
    header_row = int(structure.get("header_row", 4))
    fixed_columns = {str(column).upper() for column in structure.get("columns", {}).values() if column}
    for field_name in ("formula_columns_with_skill", "formula_columns_without_skill"):
        fixed_columns.update(
            str(column).upper()
            for column in manifest.get("fields", {}).get(field_name, {}).get("columns", [])
        )
    fixed_cells.extend(f"{column}{header_row}" for column in sorted(fixed_columns))
    regular_start = column_number(structure["columns"]["regular_items_start"])
    regular_end = column_number(structure["columns"]["regular_items_end"])
    fixed_cells.extend(
        f"{get_column_letter(column)}{header_row}"
        for column in range(regular_start, regular_end + 1)
    )
    fixed_cells = list(dict.fromkeys(str(cell) for cell in fixed_cells))
    style_row = int(structure["style_source_row"])
    format_columns = [
        structure["columns"]["student_id"],
        structure["columns"]["regular_weighted"],
        structure["columns"]["theory_weighted"],
        structure["columns"]["skill_weighted"],
        structure["columns"]["total_score"],
    ]
    return {
        "sheetnames": list(workbook.sheetnames),
        "sheet_states": {sheet.title: sheet.sheet_state for sheet in workbook.worksheets},
        "non_target_sheets": {
            sheet.title: _non_target_sheet_signature(sheet)
            for sheet in workbook.worksheets
            if sheet.title != sheet_name
        },
        "dimension": [ws.max_row, ws.max_column],
        "merged": sorted(str(item).upper() for item in ws.merged_cells.ranges),
        "page_setup": _page_setup_signature(ws),
        "worksheet_protection": _protection_signature(ws.protection),
        "workbook_protection": _protection_signature(workbook.security),
        "fixed_cells": {cell: ws[cell].value for cell in fixed_cells},
        "protected_values": _protected_target_values(ws, manifest),
        "target_cell_formats": _target_cell_format_signature(ws),
        "writable_cell_formats": {
            cell: _cell_format_signature(ws[cell]) for cell in sorted(writable_cells)
        },
        "number_formats": {f"{column}{style_row}": ws[f"{column}{style_row}"].number_format for column in format_columns},
        "orientation": ws.page_setup.orientation,
        "print_area": str(ws.print_area or ""),
        "freeze_panes": str(ws.freeze_panes or ""),
        "named_ranges": sorted(str(name) for name in workbook.defined_names),
        "data_validations": len(ws.data_validations.dataValidation),
        "conditional_formats": len(ws.conditional_formatting),
        "column_widths": {key: _dimension_signature(value.width) for key, value in ws.column_dimensions.items()},
        "row_heights": {
            str(key): _dimension_signature(value.height)
            for key, value in ws.row_dimensions.items()
            if value.height is not None
        },
    }


def _named_range_checks(
    template: Path,
    workbook,
    manifest: dict[str, Any],
    temp_dir: Path,
    report: dict[str, Any],
    errors: list[str],
) -> None:
    """Validate both the original BIFF names and the converted workbook names."""
    variant = "with_skill"
    contract = manifest["anchors"]
    xlsx_inventory = validate_named_range_inventory(workbook, contract, variant)
    report["checks"]["named_ranges_xlsx"] = xlsx_inventory
    errors.extend(xlsx_inventory["errors"])
    if template.suffix.lower() != ".xls":
        errors.append("v1.1 named-range templates must be original .xls files")
        return
    xls_inventory = validate_xls_named_range_inventory(template, contract, variant)
    report["checks"]["named_ranges_xls"] = xls_inventory
    errors.extend(xls_inventory["errors"])
    errors.extend(compare_xls_and_xlsx_named_ranges(xls_inventory, xlsx_inventory, required_names(variant)))
    errors.extend(compare_named_range_inventories(xls_inventory, xlsx_inventory, required_names(variant)))
    validate_canonical_baselines(require_v11=False)
    base_manifest = load_manifest(V10_MANIFEST)
    expected_locations = expected_named_range_locations(base_manifest["structure"], variant)
    expected_inventory = {
        "locations": {
            name: location.to_dict() for name, location in expected_locations.items()
        }
    }
    report["checks"]["expected_named_range_locations"] = expected_inventory["locations"]
    for inventory in (xls_inventory, xlsx_inventory):
        errors.extend(
            compare_named_range_inventories(
                expected_inventory,
                inventory,
                required_names(variant),
            )
        )

    locations = xlsx_inventory["locations"]
    sheet_name = locations["gb_data_table"]["sheet"] if "gb_data_table" in locations else ""
    if not sheet_name or sheet_name not in workbook.sheetnames:
        return
    ws = workbook[sheet_name]
    header_names = (
        "gb_header_serial",
        "gb_header_student_id",
        "gb_header_student_name",
        "gb_header_regular",
        "gb_header_theory",
        "gb_header_total",
    )
    expected_headers = manifest["validation"]["required_headers"]
    for name, expected in zip(header_names, expected_headers):
        location = locations.get(name)
        if location is None:
            continue
        actual = ws.cell(location["min_row"], location["min_col"]).value
        expected_text = str(expected)
        if "(" in expected_text and expected_text.endswith(")"):
            expected_text = expected_text.split("(", 1)[0]
        if expected_text not in str(actual):
            errors.append(f"Missing required header fragment {expected!r}; got {actual!r}")
    formula_names = manifest["fields"]["formula_columns_with_skill"]["names"]
    style_row = locations["gb_template_row"]["min_row"]
    for name in formula_names:
        location = locations.get(name)
        if location is None:
            continue
        cell = ws.cell(style_row, location["min_col"])
        if not isinstance(cell.value, str) or not cell.value.startswith("="):
            errors.append(f"Expected formula in named range {name} template row")
    student_id = locations.get("gb_student_id_col")
    if student_id is not None:
        if ws.cell(style_row, student_id["min_col"]).number_format != "@":
            errors.append("Student ID named column must be text formatted")
    expected_orientation = manifest["validation"].get("page_orientation", "landscape")
    if ws.page_setup.orientation != expected_orientation:
        errors.append(f"Template page orientation changed: {ws.page_setup.orientation}")
    report["checks"]["structure"] = {
        "sheets": workbook.sheetnames,
        "rows": ws.max_row,
        "columns": ws.max_column,
        "merged_count": len(ws.merged_cells.ranges),
        "expected_total_column": locations.get("gb_total_score_col", {}).get("max_col"),
        "anchor_mode": "excel_named_range",
    }

    from named_range_template_baseline import build_controlled_v11_baseline

    controlled_baseline = build_controlled_v11_baseline(
        temp_dir / "controlled-v11-baseline",
        find_soffice(),
    )
    expected_signature = _workbook_signature(
        controlled_baseline.controlled_workbook,
        base_manifest,
    )
    actual_signature = _workbook_signature(workbook, base_manifest)
    expected_signature.pop("named_ranges", None)
    actual_signature.pop("named_ranges", None)
    differences = _signature_differences(expected_signature, actual_signature)
    if differences:
        report["checks"]["protected_signature_differences"] = differences[:20]
        errors.append("Named-range template changed protected workbook structure or formatting.")


def _runtime_report(template: Path, manifest_path: Path, manifest: dict[str, Any]) -> dict[str, Any]:
    return {
        "template": str(template),
        "manifest": str(manifest_path),
        "template_version": manifest.get("template", {}).get("version"),
        "status": "failed",
        "errors": [],
        "warnings": [],
        "checks": {},
    }


def _runtime_expected_inventory() -> dict[str, Any]:
    validate_canonical_baselines(require_v11=False)
    base_manifest = load_manifest(V10_MANIFEST)
    expected_locations = expected_named_range_locations(base_manifest["structure"], "with_skill")
    return {"locations": {name: location.to_dict() for name, location in expected_locations.items()}}


def _raise_runtime_report(report: dict[str, Any], errors: list[str]) -> None:
    if errors:
        report["errors"] = sorted(set(f"Named-range runtime preflight: {error}" for error in errors))
        raise TemplateValidationError(report)


def _validate_runtime_manifest(template: Path, manifest_path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    manifest = load_manifest(manifest_path)
    ensure_supported_major(manifest)
    if validate_manifest_contract(manifest) != "excel_named_range":
        raise ValueError("Named-range runtime preflight requires a v1.1 excel_named_range manifest")
    validate_canonical_baselines(require_v11=True)
    actual_hash = validate_template_package_identity(template, manifest)
    if template.suffix.lower() != ".xls":
        raise ValueError("Named-range runtime preflight requires an original .xls template")
    return manifest, {"expected": manifest["fingerprint"]["sha256"], "actual": actual_hash}


def validate_named_range_runtime_raw(
    template_path: Path | str,
    manifest_path: Path | str = DEFAULT_MANIFEST,
) -> dict[str, Any]:
    """Run the non-skippable BIFF-only name safety contract."""
    template = Path(template_path).expanduser().resolve()
    manifest_file = Path(manifest_path).expanduser().resolve()
    manifest, fingerprint = _validate_runtime_manifest(template, manifest_file)
    report = _runtime_report(template, manifest_file, manifest)
    report["checks"]["sha256"] = fingerprint
    inventory = validate_xls_named_range_inventory(template, manifest["anchors"], "with_skill")
    expected_inventory = _runtime_expected_inventory()
    report["checks"]["named_ranges_xls"] = inventory
    report["checks"]["expected_named_range_locations"] = expected_inventory["locations"]
    errors = list(inventory["errors"])
    errors.extend(compare_named_range_inventories(expected_inventory, inventory, required_names("with_skill")))
    report["checks"]["named_range_count"] = len(required_names("with_skill"))
    report["checks"]["runtime_contract"] = {
        "scope": "workbook",
        "regular_item_count": manifest["validation"]["regular_item_count"],
        "regular_items_width": 8,
        "required_variant": "with_skill",
        "forbidden_variant_names": [],
    }
    _raise_runtime_report(report, errors)
    report["status"] = "passed"
    return report


def validate_named_range_runtime_roundtrip(
    template_path: Path | str,
    manifest_path: Path | str = DEFAULT_MANIFEST,
    raw_report: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Run the optional LibreOffice XLS->XLSX consistency checks."""
    template = Path(template_path).expanduser().resolve()
    manifest_file = Path(manifest_path).expanduser().resolve()
    if raw_report is None:
        raw_report = validate_named_range_runtime_raw(template, manifest_file)
    manifest, fingerprint = _validate_runtime_manifest(template, manifest_file)
    report = _runtime_report(template, manifest_file, manifest)
    report["checks"]["sha256"] = fingerprint
    report["checks"]["named_ranges_xls"] = raw_report["checks"]["named_ranges_xls"]
    expected_inventory = _runtime_expected_inventory()
    report["checks"]["expected_named_range_locations"] = expected_inventory["locations"]
    errors: list[str] = []
    with tempfile.TemporaryDirectory(prefix="gradebook-named-runtime-roundtrip-") as temp_name:
        xlsx = convert_to_xlsx(template, Path(temp_name) / "roundtrip-xlsx", find_soffice())
        workbook = load_workbook(xlsx, data_only=False)
        xlsx_inventory = validate_named_range_inventory(workbook, manifest["anchors"], "with_skill")
        report["checks"]["named_ranges_xlsx"] = xlsx_inventory
        errors.extend(xlsx_inventory["errors"])
        errors.extend(compare_xls_and_xlsx_named_ranges(
            raw_report["checks"]["named_ranges_xls"], xlsx_inventory, required_names("with_skill")
        ))
        errors.extend(compare_named_range_inventories(
            raw_report["checks"]["named_ranges_xls"], xlsx_inventory, required_names("with_skill")
        ))
        errors.extend(compare_named_range_inventories(expected_inventory, xlsx_inventory, required_names("with_skill")))
    report["checks"]["named_range_count"] = len(required_names("with_skill"))
    _raise_runtime_report(report, errors)
    report["status"] = "passed"
    return report


def validate_named_range_runtime(
    template_path: Path | str,
    manifest_path: Path | str = DEFAULT_MANIFEST,
) -> dict[str, Any]:
    """Run raw and LibreOffice round-trip runtime checks."""
    raw_report = validate_named_range_runtime_raw(template_path, manifest_path)
    return validate_named_range_runtime_roundtrip(template_path, manifest_path, raw_report)


def validate_template(
    template_path: Path | str,
    manifest_path: Path | str = DEFAULT_MANIFEST,
    compatibility_template: Path | str | None = None,
) -> dict[str, Any]:
    template = Path(template_path).expanduser().resolve()
    manifest = load_manifest(manifest_path)
    errors: list[str] = []
    warnings: list[str] = []
    report: dict[str, Any] = {
        "template": str(template),
        "manifest": str(Path(manifest_path).expanduser().resolve()),
        "template_version": manifest.get("template", {}).get("version"),
        "errors": errors,
        "warnings": warnings,
        "checks": {},
    }
    try:
        ensure_supported_major(manifest)
    except ValueError as exc:
        errors.append(str(exc))
    try:
        mode = validate_manifest_contract(manifest)
    except ValueError as exc:
        mode = ""
        errors.append(str(exc))
    if not template.exists():
        errors.append(f"Template not found: {template}")
        raise TemplateValidationError(report)

    expected_hash = str(manifest.get("fingerprint", {}).get("sha256") or manifest.get("fingerprint", {}).get("value", "")).upper()
    actual_hash = sha256(template)
    try:
        canonical = canonical_template_for_mode(mode)
        validate_canonical_baselines(require_v11=mode == "excel_named_range")
    except ValueError as exc:
        errors.append(str(exc))
        canonical = manifest_template_path(manifest)
    is_canonical = template == canonical
    if not is_canonical and not canonical.exists():
        errors.append(f"Canonical template not found: {canonical}")
    report["checks"]["sha256"] = {"expected": expected_hash, "actual": actual_hash}
    if actual_hash != expected_hash:
        errors.append(f"Template fingerprint mismatch: expected {expected_hash}, got {actual_hash}")

    if compatibility_template is None:
        entries = manifest.get("template", {}).get("compatibility_entries", [])
        if entries:
            compatibility_template = Path(manifest["_path"]).parent / entries[0]
    if compatibility_template:
        compat = Path(compatibility_template).expanduser().resolve()
        if compat.exists():
            compat_hash = sha256(compat)
            report["checks"]["compatibility_sha256"] = {"path": str(compat), "actual": compat_hash}
            if compat_hash != expected_hash:
                errors.append(f"Compatibility template diverges from canonical template: {compat}")
        else:
            warnings.append(f"Compatibility template entry is missing: {compat}")

    if mode == "excel_named_range":
        try:
            with tempfile.TemporaryDirectory(prefix="gradebook-named-template-") as temp_name:
                xlsx = convert_to_xlsx(template, Path(temp_name), find_soffice())
                workbook = load_workbook(xlsx, data_only=False)
                report["checks"]["workbook_open"] = True
                _named_range_checks(template, workbook, manifest, Path(temp_name), report, errors)
        except TemplateValidationError:
            raise
        except Exception as exc:
            errors.append(f"XLS named-range template could not be opened or inspected: {exc}")
        if errors:
            raise TemplateValidationError(report)
        return report

    try:
        with tempfile.TemporaryDirectory(prefix="gradebook-template-") as temp_name:
            xlsx = template if template.suffix.lower() == ".xlsx" else convert_to_xlsx(template, Path(temp_name), find_soffice())
            workbook = load_workbook(xlsx, data_only=False)
            report["checks"]["workbook_open"] = True
            structure = manifest["structure"]
            sheet_name = structure["worksheet"]
            if sheet_name not in workbook.sheetnames:
                errors.append(f"Missing worksheet: {sheet_name}")
                raise TemplateValidationError(report)
            ws = workbook[sheet_name]
            expected_sheets = list(structure.get("required_sheets", []))
            if expected_sheets and workbook.sheetnames != expected_sheets:
                errors.append(f"Worksheet names changed: expected {expected_sheets}, got {workbook.sheetnames}")
            expected_states = {str(key): str(value) for key, value in structure.get("required_sheet_states", {}).items()}
            actual_states = {sheet.title: sheet.sheet_state for sheet in workbook.worksheets}
            if expected_states and actual_states != expected_states:
                errors.append(f"Worksheet visibility changed: expected {expected_states}, got {actual_states}")
            expected_last_row = int(structure["template_last_data_row"])
            if ws.max_row != expected_last_row:
                errors.append(f"Template row count mismatch: expected {expected_last_row}, got {ws.max_row}")
            expected_total_col = structure["columns"]["total_score"]
            expected_columns = column_number(structure["columns"]["total_score"])
            if ws.max_column != expected_columns:
                errors.append(f"Template column count mismatch: expected {expected_columns}, got {ws.max_column}")
            merged = set(str(item).upper() for item in ws.merged_cells.ranges)
            required_merged = set(str(item).upper() for item in structure.get("required_merged_ranges", []))
            if merged != required_merged:
                errors.append(f"Protected merged ranges changed: expected {sorted(required_merged)}, got {sorted(merged)}")
            required_headers = manifest["validation"]["required_headers"]
            label_cells = structure.get("header_label_cells", {"serial": "A3", "student_id": "B3", "student_name": "C3", "regular": "D3", "theory": "M3", "total": "Q3"})
            header_values = [ws[label_cells[key]].value for key in ("serial", "student_id", "student_name", "regular", "theory", "total")]
            for expected, actual in zip(required_headers, header_values):
                expected_text = str(expected)
                if not is_canonical and "(" in expected_text and expected_text.endswith(")"):
                    expected_text = expected_text.split("(", 1)[0]
                if expected_text not in str(actual):
                    errors.append(f"Missing required header fragment {expected!r}; got {actual!r}")
            formula_columns = manifest["fields"].get("formula_columns_with_skill", {}).get("columns", ["L", "N", "P", "Q"])
            formula_row = int(structure["style_source_row"])
            for column in formula_columns:
                cell = f"{column}{formula_row}"
                if not isinstance(ws[cell].value, str) or not ws[cell].value.startswith("="):
                    errors.append(f"Expected formula in template cell {cell}")
            student_id_column = structure["columns"]["student_id"]
            student_id_cell = f"{student_id_column}{formula_row}"
            if ws[student_id_cell].number_format != "@":
                errors.append(f"Student ID cell {student_id_cell} must be text formatted, got {ws[student_id_cell].number_format!r}")
            expected_orientation = manifest["validation"].get("page_orientation", "landscape")
            if ws.page_setup.orientation != expected_orientation:
                errors.append(f"Template page orientation changed: {ws.page_setup.orientation}")
            expected_print_area = str(manifest["validation"].get("expected_print_area") or "")
            if str(ws.print_area or "") != expected_print_area:
                errors.append(f"Template print area changed: expected {expected_print_area!r}, got {str(ws.print_area or '')!r}")
            expected_freeze = str(manifest["validation"].get("expected_freeze_panes") or "")
            if str(ws.freeze_panes or "") != expected_freeze:
                errors.append(f"Template freeze panes changed: expected {expected_freeze!r}, got {str(ws.freeze_panes or '')!r}")
            expected_named_ranges = sorted(str(item) for item in manifest["validation"].get("required_named_ranges", []))
            actual_named_ranges = sorted(str(name) for name in workbook.defined_names)
            if actual_named_ranges != expected_named_ranges:
                errors.append(f"Named ranges changed: expected {expected_named_ranges}, got {actual_named_ranges}")
            expected_dv = int(manifest["validation"].get("required_data_validations", 0))
            actual_dv = len(ws.data_validations.dataValidation)
            if actual_dv != expected_dv:
                errors.append(f"Data validations changed: expected {expected_dv}, got {actual_dv}")
            expected_cf = int(manifest["validation"].get("required_conditional_formats", 0))
            actual_cf = len(ws.conditional_formatting)
            if actual_cf != expected_cf:
                errors.append(f"Conditional formats changed: expected {expected_cf}, got {actual_cf}")
            report["checks"]["structure"] = {
                "sheets": workbook.sheetnames,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "merged_count": len(merged),
                "expected_total_column": expected_total_col,
            }
            if canonical.exists() and not is_canonical:
                canonical_xlsx = canonical
                with tempfile.TemporaryDirectory(prefix="gradebook-canonical-") as canonical_temp:
                    canonical_xlsx = convert_to_xlsx(canonical, Path(canonical_temp), find_soffice())
                    canonical_workbook = load_workbook(canonical_xlsx, data_only=False)
                    canonical_signature = _workbook_signature(canonical_workbook, manifest)
                    custom_signature = _workbook_signature(workbook, manifest)
                    controlled_signature = None
                    raw_font_differences: list[dict[str, Any]] = []
                    if template.suffix.lower() == ".xls":
                        static_controlled_xls, static_controlled_xlsx = controlled_roundtrip_paths(
                            canonical,
                            Path(canonical_temp) / "static-controlled",
                            find_soffice(),
                        )
                        static_controlled_workbook = load_workbook(static_controlled_xlsx, data_only=False)
                        static_controlled_signature = _workbook_signature(static_controlled_workbook, manifest)
                        content_controlled_xls, content_controlled_xlsx = controlled_content_roundtrip_paths(
                            canonical_xlsx,
                            xlsx,
                            sheet_name,
                            Path(canonical_temp) / "controlled",
                            find_soffice(),
                        )
                        content_controlled_workbook = load_workbook(content_controlled_xlsx, data_only=False)
                        content_controlled_signature = _workbook_signature(content_controlled_workbook, manifest)
                        controlled_signature = dict(static_controlled_signature)
                        controlled_signature["target_cell_formats"] = dict(
                            static_controlled_signature["target_cell_formats"]
                        )
                        controlled_signature["writable_cell_formats"] = dict(
                            static_controlled_signature["writable_cell_formats"]
                        )
                        canonical_ws = canonical_workbook[sheet_name]
                        custom_ws = workbook[sheet_name]
                        writable_cells = {
                            str(cell)
                            for cell in [
                                *manifest["structure"].get("metadata", {}).values(),
                                *manifest["structure"].get("headers", {}).values(),
                            ]
                            if cell
                        }
                        changed_writable_anchors: set[str] = set()
                        for address in writable_cells:
                            if canonical_ws[address].value != custom_ws[address].value:
                                changed_writable_anchors.add(address)
                                for section_name in ("target_cell_formats", "writable_cell_formats"):
                                    controlled_signature[section_name][address] = (
                                        content_controlled_signature[section_name].get(address)
                                    )
                        changed_writable_cells = set(changed_writable_anchors)
                        for merged in canonical_ws.merged_cells.ranges:
                            if merged.start_cell.coordinate not in changed_writable_anchors:
                                continue
                            changed_writable_cells.update(
                                f"{get_column_letter(column)}{row}"
                                for row, column in merged.cells
                            )
                        for address in changed_writable_cells - changed_writable_anchors:
                            controlled_signature["target_cell_formats"][address] = (
                                content_controlled_signature["target_cell_formats"].get(address)
                            )
                        unchanged_xls = canonical if actual_hash == expected_hash else static_controlled_xls
                        for section_name in ("target_cell_formats", "writable_cell_formats"):
                            for address in canonical_signature.get(section_name, {}):
                                expected_xls = (
                                    content_controlled_xls
                                    if address in changed_writable_cells
                                    else unchanged_xls
                                )
                                expected_font = xls_font_identity(expected_xls, sheet_name, address)
                                actual_font = xls_font_identity(template, sheet_name, address)
                                if actual_font is not None and expected_font is not None and actual_font != expected_font:
                                    raw_font_differences.append(
                                        {
                                            "path": f"{section_name}.{address}.font",
                                            "expected": expected_font,
                                            "actual": actual_font,
                                        }
                                    )
                    differences = _signature_differences(
                        canonical_signature,
                        custom_signature,
                        font_matcher=lambda expected, actual, path: _controlled_font_fallback_at_path(
                            expected,
                            actual,
                            path,
                            controlled_signature,
                        ),
                    )
                    differences = raw_font_differences + differences
                    if differences:
                        report["checks"]["protected_signature_differences"] = differences[:20]
                        errors.append("Custom template changed protected workbook structure or formatting.")
    except TemplateValidationError:
        raise
    except Exception as exc:
        errors.append(f"XLS template could not be opened or inspected: {exc}")

    if errors:
        raise TemplateValidationError(report)
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate the versioned XLS course-gradebook template package.")
    parser.add_argument("--template", default="")
    parser.add_argument("--manifest", default="")
    parser.add_argument("--compatibility-template", default="")
    parser.add_argument("--identity-only", action="store_true", help="Only verify the template fingerprint and manifest contract.")
    parser.add_argument(
        "--named-range-runtime-preflight",
        action="store_true",
        help="Run the non-skippable v1.1 named-range safety checks without full formatting validation.",
    )
    parser.add_argument(
        "--named-range-runtime-raw",
        action="store_true",
        help="Run the non-skippable raw BIFF named-range safety checks without LibreOffice.",
    )
    parser.add_argument("--json", action="store_true", dest="as_json")
    args = parser.parse_args()
    try:
        package = resolve_template_package(args.template or None, args.manifest or None)
        manifest = package.manifest
        if args.named_range_runtime_raw:
            report = validate_named_range_runtime_raw(package.template_path, package.manifest_path)
        elif args.named_range_runtime_preflight:
            report = validate_named_range_runtime(package.template_path, package.manifest_path)
        elif args.identity_only:
            actual = validate_template_package_identity(package.template_path, manifest)
            report = {
                "template": str(package.template_path),
                "manifest": str(package.manifest_path),
                "template_version": manifest.get("template", {}).get("version"),
                "errors": [],
                "warnings": [],
                "checks": {"sha256": {"expected": manifest.get("fingerprint", {}).get("sha256"), "actual": actual}},
            }
        else:
            report = validate_template(package.template_path, package.manifest_path, args.compatibility_template or None)
    except TemplateValidationError as exc:
        report = exc.report
        if args.as_json:
            print(json.dumps(report, ensure_ascii=False, indent=2))
        else:
            for error in report.get("errors", []):
                print(f"ERROR: {error}", file=sys.stderr)
            for warning in report.get("warnings", []):
                print(f"WARNING: {warning}", file=sys.stderr)
            differences = report.get("checks", {}).get("protected_signature_differences", [])
            if differences:
                print("PROTECTED_SIGNATURE_DIFFERENCES:", file=sys.stderr)
                for difference in differences[:10]:
                    print(
                        "  "
                        + json.dumps(difference, ensure_ascii=False, sort_keys=True),
                        file=sys.stderr,
                    )
        return 1
    except Exception as exc:
        report = {
            "template": str(Path(args.template).expanduser().resolve()) if args.template else "",
            "manifest": str(Path(args.manifest).expanduser().resolve()) if args.manifest else "",
            "template_version": None,
            "errors": [str(exc)],
            "warnings": [],
            "checks": {},
        }
        if args.as_json:
            print(json.dumps(report, ensure_ascii=False, indent=2))
        else:
            print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    if args.as_json:
        print(json.dumps(report, ensure_ascii=False, indent=2))
    else:
        print(
            f"validated template={package.template_path} version={report['template_version']} "
            f"sha256={report['checks']['sha256']['actual']}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
