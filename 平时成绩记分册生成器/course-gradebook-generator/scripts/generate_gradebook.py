#!/usr/bin/env python3
"""Cross-platform gradebook generator for 课程成绩单.xls -> 平时成绩记分册.xls.

Requires Python 3, openpyxl, and LibreOffice/soffice on PATH.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import platform
import random
import re
import shutil
import subprocess
import sys
import tempfile
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from named_range_contracts import variant_for_skill
from named_range_utils import (
    NamedRangeError,
    rebuild_named_ranges_after_column_delete,
    update_named_ranges_for_capacity,
    validate_named_range_inventory,
    variant_locations,
)
from xls_named_range_utils import validate_xls_named_range_inventory
from package_common import (
    DEFAULT_MANIFEST,
    DEFAULT_SCHEMA,
    anchor_mode,
    cell_address,
    column_number,
    ensure_supported_major,
    load_manifest,
    manifest_template_path,
    percentage_label,
    resolve_template_package,
    validate_manifest_contract,
    validate_template_package_identity,
    validate_input,
    validate_source_totals,
)
from validate_output import validate_output_dir, write_skipped_report
from validate_template import validate_template


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
DEFAULT_TEMPLATE = SKILL_DIR / "assets" / "templates" / "course-gradebook" / "v1.1.0" / "template.xls"


def find_soffice() -> str:
    candidates = [
        shutil.which("soffice"),
        shutil.which("libreoffice"),
        shutil.which("soffice.com"),
        r"C:\Program Files\LibreOffice\program\soffice.com",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/usr/bin/libreoffice",
        "/usr/local/bin/libreoffice",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(candidate)
    raise RuntimeError(
        "LibreOffice/soffice was not found. Install LibreOffice and make soffice available on PATH."
    )


def convert_with_soffice(soffice: str, source: Path, out_dir: Path, target: str) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    cmd = [soffice, "--headless", "--convert-to", target, "--outdir", str(out_dir), str(source)]
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, encoding="utf-8", errors="replace")
    if proc.returncode != 0:
        raise RuntimeError(f"LibreOffice conversion failed: {' '.join(cmd)}\n{proc.stdout}\n{proc.stderr}")
    ext = target.split(":")[0].lower()
    output = out_dir / f"{source.stem}.{ext}"
    if not output.exists():
        matches = sorted(out_dir.glob(f"{source.stem}.*"))
        if matches:
            return matches[-1]
        raise RuntimeError(f"LibreOffice did not create expected output for {source}")
    return output


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_number(value) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


def read_meta(ws, manifest: dict | None = None) -> dict:
    source = (manifest or {}).get("structure", {}).get("source", {})
    line2_cell = str(source.get("metadata_line2_cell", "A2"))
    line3_cell = str(source.get("metadata_line3_cell", "A3"))
    line2 = cell_text(ws[line2_cell].value)
    line3 = cell_text(ws[line3_cell].value)
    meta = {
        "course": "",
        "teacher": "",
        "class_name": "",
        "term": "",
        "skill_pct": 0.0,
        "theory_pct": 0.0,
        "regular_pct": 0.0,
    }
    m = re.search(r"课程名称:([^\r\n]+?)\s+教师:", line2)
    if not m:
        m = re.search(r"课程名称:([^\r\n ]+)", line2)
    if m:
        meta["course"] = m.group(1).strip()
    m = re.search(r"教师:([^\r\n]+?)(?:\s*上课班级:|$)", line2)
    if m:
        meta["teacher"] = m.group(1).strip()
    m = re.search(r"上课班级:([^\r\n]+?)(?:\s*成绩项目比例:|$)", line2)
    if m:
        meta["class_name"] = m.group(1).strip()
    m = re.search(r"开课学期:([^\s]+)", line3)
    if m:
        meta["term"] = m.group(1).strip()
    m = re.search(r"技能成绩(\d+(?:\.\d+)?)%", line2)
    if m:
        meta["skill_pct"] = float(m.group(1)) / 100
    m = re.search(r"理论成绩(\d+(?:\.\d+)?)%", line2)
    if m:
        meta["theory_pct"] = float(m.group(1)) / 100
    m = re.search(r"平时成绩(\d+(?:\.\d+)?)%", line2)
    if m:
        meta["regular_pct"] = float(m.group(1)) / 100
    return meta


def header_map(ws, start_col: int, end_col: int, header_row: int = 4) -> dict[str, int]:
    found = {}
    for col in range(start_col, end_col + 1):
        header = cell_text(ws.cell(header_row, col).value)
        if header:
            found[header] = col
    return found


def read_students(ws, manifest: dict | None = None) -> list[dict]:
    source = (manifest or {}).get("structure", {}).get("source", {})
    header_row = int(source.get("header_row", 4))
    data_start_row = int(source.get("data_start_row", 5))
    header_names = {
        "student_id": "学号",
        "student_name": "姓名",
        "regular": "平时成绩",
        "theory": "理论成绩",
        "skill": "技能成绩",
        "total": "总成绩",
        **source.get("headers", {}),
    }
    starts = []
    for col in range(1, ws.max_column + 1):
        if cell_text(ws.cell(header_row, col).value) == header_names["student_id"] and cell_text(ws.cell(header_row, col + 1).value) == header_names["student_name"]:
            starts.append(col)
    if not starts:
        raise RuntimeError("Could not find 学号/姓名 headers in source workbook.")

    blocks = []
    for i, start in enumerate(starts):
        end = starts[i + 1] - 1 if i + 1 < len(starts) else ws.max_column
        blocks.append((start, header_map(ws, start, end, header_row)))

    students = []
    for row in range(data_start_row, ws.max_row + 1):
        for start, headers in blocks:
            student_id = cell_text(ws.cell(row, start).value)
            name = cell_text(ws.cell(row, start + 1).value)
            if not re.fullmatch(r"\d{8,}", student_id) or not name:
                continue
            for required in (header_names["theory"], header_names["regular"], header_names["total"]):
                if required not in headers:
                    raise RuntimeError(f"Source block is missing {required}.")
            students.append(
                {
                    "id": student_id,
                    "name": name,
                    "skill": to_number(ws.cell(row, headers[header_names["skill"]]).value) if header_names["skill"] in headers else 0.0,
                    "theory": to_number(ws.cell(row, headers[header_names["theory"]]).value),
                    "regular": to_number(ws.cell(row, headers[header_names["regular"]]).value),
                    "total": to_number(ws.cell(row, headers[header_names["total"]]).value),
                }
            )
    if not students:
        raise RuntimeError("No students parsed from source workbook.")
    return students


def stable_seed(text: str) -> int:
    digest = hashlib.sha256(text.encode("utf-8")).digest()
    return int.from_bytes(digest[:4], "little") & 0x7FFFFFFF


def generate_regular_scores(target: float, seed_text: str, item_count: int = 8) -> list[float]:
    if item_count <= 0:
        raise ValueError("regular item count must be positive")
    target_units = round(target * 2)
    target_total_units = target_units * item_count
    rng = random.Random(stable_seed(seed_text))
    for _ in range(2000):
        scores = []
        total = 0
        for _idx in range(item_count - 1):
            low = max(-8, -target_units)
            high = min(8, 200 - target_units)
            score = target_units + rng.randint(low, high)
            scores.append(score)
            total += score
        last = target_total_units - total
        if last < 0 or last > 200:
            continue
        if abs(last - target_units) > 12:
            continue
        scores.append(last)
        values = [score / 2 for score in scores]
        if max(abs(value - target) for value in values) <= 6:
            return values
    return [target] * item_count


def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)


def writable_cell(ws, row: int, col: int):
    coord = f"{get_column_letter(col)}{row}"
    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            cell = ws.cell(merged_range.min_row, merged_range.min_col)
            if isinstance(cell, MergedCell):
                ws.unmerge_cells(str(merged_range))
                return ws.cell(row, col)
            return cell
    return ws.cell(row, col)


def set_value(ws, row: int, col: int, value) -> None:
    writable_cell(ws, row, col).value = value


def _shift_column_after_delete(column: int, start: int, count: int) -> int | None:
    if column < start:
        return column
    if column >= start + count:
        return column - count
    return None


def delete_columns_preserving_merges(ws, start_col: int, count: int) -> None:
    """Delete optional columns while retaining every surviving merged range."""
    original_ranges = [
        (merged.min_row, merged.max_row, merged.min_col, merged.max_col)
        for merged in list(ws.merged_cells.ranges)
    ]
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    ws.delete_cols(start_col, count)
    for min_row, max_row, min_col, max_col in original_ranges:
        shifted_min = _shift_column_after_delete(min_col, start_col, count)
        shifted_max = _shift_column_after_delete(max_col, start_col, count)
        if shifted_min is None:
            shifted_min = start_col
        if shifted_max is None:
            shifted_max = start_col - 1
        if min_col < start_col and max_col >= start_col + count:
            shifted_min = min_col
            shifted_max = max_col - count
        elif min_col < start_col <= max_col:
            shifted_min = min_col
            shifted_max = start_col - 1
        elif min_col >= start_col and max_col < start_col + count:
            continue
        if shifted_max < shifted_min:
            continue
        if shifted_min != shifted_max or min_row != max_row:
            ws.merge_cells(
                start_row=min_row,
                start_column=shifted_min,
                end_row=max_row,
                end_column=shifted_max,
            )


def formula_number(value: float) -> str:
    return f"{value:.8f}".rstrip("0").rstrip(".") or "0"


def _set_address(ws, address: str, value) -> None:
    column = re.match(r"([A-Z]+)(\d+)$", address, re.IGNORECASE)
    if not column:
        raise ValueError(f"Invalid manifest cell address: {address}")
    set_value(ws, int(column.group(2)), column_number(column.group(1)), value)


def build_one_legacy(source_xls: Path, template_xls: Path, output_dir: Path, soffice: str, manifest: dict, schema_path: Path) -> dict:
    with tempfile.TemporaryDirectory(prefix="gradebook_") as tmp_name:
        tmp = Path(tmp_name)
        source_xlsx = convert_with_soffice(soffice, source_xls, tmp / "source", "xlsx")
        template_xlsx = convert_with_soffice(soffice, template_xls, tmp / "template", "xlsx")

        src_wb = load_workbook(source_xlsx, data_only=True)
        src_ws = src_wb.worksheets[0]
        meta = read_meta(src_ws, manifest)
        students = read_students(src_ws, manifest)

        normalized = {
            "term": meta["term"],
            "course": meta["course"],
            "teacher": meta["teacher"],
            "class_name": meta["class_name"],
            "weights": {
                "regular": meta["regular_pct"],
                "theory": meta["theory_pct"],
                "skill": meta["skill_pct"],
            },
            "students": students,
        }
        validate_input(normalized, schema_path)
        validate_source_totals(students, normalized["weights"])

        wb = load_workbook(template_xlsx)
        structure = manifest["structure"]
        columns = structure["columns"]
        ws = wb[structure["worksheet"]] if structure["worksheet"] in wb.sheetnames else wb.worksheets[0]
        has_skill = meta["skill_pct"] > 0.000001
        skill_start = column_number(columns["skill_score"])
        class_name_cell = structure["metadata"]["class_name"]
        class_name_style = copy(ws[class_name_cell]._style)
        if not has_skill:
            delete_columns_preserving_merges(ws, skill_start, 2)
            # Deleting the optional columns removes the original O2:Q2 anchor;
            # restore the surviving class-name cell's protected base style.
            ws[class_name_cell]._style = copy(class_name_style)

        data_start = int(structure["data_start_row"])
        template_last_row = int(structure["template_last_data_row"])
        style_source_row = int(structure["style_source_row"])
        regular_item_count = int(manifest["validation"]["regular_item_count"])
        if regular_item_count <= 0:
            raise ValueError("Manifest validation.regular_item_count must be positive")
        regular_start = column_number(columns["regular_items_start"])
        regular_end = column_number(columns["regular_items_end"])
        if regular_end - regular_start + 1 != regular_item_count:
            raise ValueError(
                "Manifest regular item count does not match columns.regular_items_start/regular_items_end"
            )
        total_col = column_number(columns["total_score"]) if has_skill else column_number(structure["no_skill_total_column"])
        max_col = total_col
        existing_rows = template_last_row - data_start + 1
        if len(students) > existing_rows:
            needed = len(students) - existing_rows
            ws.insert_rows(template_last_row + 1, needed)
            for row in range(template_last_row + 1, template_last_row + needed + 1):
                copy_row_style(ws, style_source_row, row, max_col)
            template_last_row += needed

        for row in range(data_start, template_last_row + 1):
            for col in range(1, max_col + 1):
                ws.cell(row, col).value = None

        for name, value in {
            "term": meta["term"],
            "course": meta["course"],
            "teacher": meta["teacher"],
            "class_name": meta["class_name"],
        }.items():
            _set_address(ws, structure["metadata"][name], value)
        _set_address(ws, structure["headers"]["regular"], f"平时成绩({percentage_label(meta['regular_pct'])}%)")
        _set_address(ws, structure["headers"]["theory"], f"理论成绩({percentage_label(meta['theory_pct'])}%)")
        if has_skill:
            _set_address(ws, structure["headers"]["skill"], f"技能成绩（{percentage_label(meta['skill_pct'])}%）")
        else:
            no_skill_total = structure["no_skill_total_column"]
            ws.column_dimensions[no_skill_total].width = max(ws.column_dimensions[no_skill_total].width or 0, 18)

        regular_pct = formula_number(meta["regular_pct"])
        theory_pct = formula_number(meta["theory_pct"])
        skill_pct = formula_number(meta["skill_pct"])
        class_code = source_xls.parent.name or source_xls.stem

        for idx, student in enumerate(students):
            row = data_start + idx
            scores = generate_regular_scores(
                student["regular"], f"{class_code}|{student['id']}|{student['regular']}", regular_item_count
            )
            ws.cell(row, column_number(columns["serial"])).value = idx + 1
            ws.cell(row, column_number(columns["student_id"])).value = student["id"]
            ws.cell(row, column_number(columns["student_id"])).number_format = "@"
            ws.cell(row, column_number(columns["student_name"])).value = student["name"]
            for offset, score in enumerate(scores):
                ws.cell(row, regular_start + offset).value = score
            ws.cell(row, column_number(columns["regular_weighted"])).value = f"=AVERAGE({columns['regular_items_start']}{row}:{columns['regular_items_end']}{row})*{regular_pct}"
            ws.cell(row, column_number(columns["theory_score"])).value = student["theory"]
            ws.cell(row, column_number(columns["theory_weighted"])).value = f"={columns['theory_score']}{row}*{theory_pct}"
            if has_skill:
                ws.cell(row, column_number(columns["skill_score"])).value = student["skill"]
                ws.cell(row, column_number(columns["skill_weighted"])).value = f"={columns['skill_score']}{row}*{skill_pct}"
                ws.cell(row, total_col).value = f"=ROUND(AVERAGE({columns['regular_items_start']}{row}:{columns['regular_items_end']}{row})*{regular_pct}+{columns['theory_score']}{row}*{theory_pct}+{columns['skill_score']}{row}*{skill_pct},0)"
            else:
                ws.cell(row, total_col).value = f"=ROUND(AVERAGE({columns['regular_items_start']}{row}:{columns['regular_items_end']}{row})*{regular_pct}+{columns['theory_score']}{row}*{theory_pct},0)"

        # Keep exactly as many student rows as the source contains.
        first_extra = data_start + len(students)
        if first_extra <= template_last_row:
            ws.delete_rows(first_extra, template_last_row - first_extra + 1)

        output_dir.mkdir(parents=True, exist_ok=True)
        temp_xlsx = tmp / f"{class_code}-平时成绩记分册.xlsx"
        wb.save(temp_xlsx)
        converted = convert_with_soffice(soffice, temp_xlsx, output_dir, "xls")
        final_path = output_dir / f"{class_code}-平时成绩记分册.xls"
        if converted.resolve() != final_path.resolve():
            if final_path.exists():
                final_path.unlink()
            converted.replace(final_path)

        return {
            "source": str(source_xls),
            "output": str(final_path),
            "count": len(students),
            "course": meta["course"],
            "class_name": meta["class_name"],
            "has_skill": has_skill,
            "regular_pct": meta["regular_pct"],
            "theory_pct": meta["theory_pct"],
            "skill_pct": meta["skill_pct"],
            "engine": "libreoffice-openpyxl",
            "platform": platform.system(),
            "normalized_input": normalized,
        }


def _named_cell(ws, location: dict[str, int]):
    return ws.cell(location["min_row"], location["min_col"])


def _named_output_cell(workbook, locations: dict, name: str, row_offset: int = 0, col_offset: int = 0):
    location = locations[name]
    row = location.min_row + row_offset
    col = location.min_col + col_offset
    if row > location.max_row or col > location.max_col:
        raise NamedRangeError(
            f"Write offset {row_offset},{col_offset} is outside managed name {name} at {location.address}"
        )
    return workbook[location.sheet].cell(row, col)


def _clear_named_output_range(workbook, locations: dict, name: str) -> None:
    location = locations[name]
    sheet = workbook[location.sheet]
    for row in range(location.min_row, location.max_row + 1):
        for col in range(location.min_col, location.max_col + 1):
            sheet.cell(row, col).value = None


def _named_col_letter(location: dict[str, int]) -> str:
    if location["min_col"] != location["max_col"]:
        raise NamedRangeError(f"Expected a single named column, got {location['address']}")
    return get_column_letter(location["min_col"])


def build_one_named_ranges(
    source_xls: Path,
    template_xls: Path,
    output_dir: Path,
    soffice: str,
    manifest: dict,
    schema_path: Path,
) -> dict:
    """Build v1.1 output exclusively from the workbook-level name contract."""
    with tempfile.TemporaryDirectory(prefix="gradebook-named-") as tmp_name:
        tmp = Path(tmp_name)
        source_xlsx = convert_with_soffice(soffice, source_xls, tmp / "source", "xlsx")
        template_xlsx = convert_with_soffice(soffice, template_xls, tmp / "template", "xlsx")

        src_wb = load_workbook(source_xlsx, data_only=True)
        src_ws = src_wb.worksheets[0]
        meta = read_meta(src_ws, manifest)
        students = read_students(src_ws, manifest)
        normalized = {
            "term": meta["term"],
            "course": meta["course"],
            "teacher": meta["teacher"],
            "class_name": meta["class_name"],
            "weights": {
                "regular": meta["regular_pct"],
                "theory": meta["theory_pct"],
                "skill": meta["skill_pct"],
            },
            "students": students,
        }
        validate_input(normalized, schema_path)
        validate_source_totals(students, normalized["weights"])

        workbook = load_workbook(template_xlsx, data_only=False)
        template_inventory = validate_named_range_inventory(workbook, manifest["anchors"], "with_skill")
        if template_inventory["errors"]:
            raise NamedRangeError("v1.1 template named-range inventory is invalid: " + "; ".join(template_inventory["errors"]))
        has_skill = meta["skill_pct"] > 0.000001
        variant = variant_for_skill(has_skill)
        locations = variant_locations(workbook, "with_skill")
        table = locations["gb_data_table"]
        sheet_name = table.sheet
        ws = workbook[sheet_name]
        class_name_cell = _named_cell(ws, locations["gb_class_name"].to_dict())
        class_name_style = copy(class_name_cell._style)
        if not has_skill:
            skill_start = locations["gb_skill_score_col"].min_col
            delete_columns_preserving_merges(ws, skill_start, 2)
            ws.cell(locations["gb_class_name"].min_row, skill_start)._style = copy(class_name_style)
            rebuild_named_ranges_after_column_delete(workbook, skill_start, 2, variant)
        locations = variant_locations(workbook, variant)
        table = locations["gb_data_table"]
        data_start = table.min_row
        template_last_row = table.max_row
        style_source_row = locations["gb_template_row"].min_row
        total_col = locations["gb_total_score_col"].min_col
        regular_items = locations["gb_regular_items"]
        regular_item_count = int(manifest["validation"]["regular_item_count"])
        if regular_items.width != regular_item_count:
            raise NamedRangeError(
                f"Named range gb_regular_items has {regular_items.width} columns; expected {regular_item_count}"
            )
        existing_rows = template_last_row - data_start + 1
        if len(students) > existing_rows:
            needed = len(students) - existing_rows
            ws.insert_rows(template_last_row + 1, needed)
            for row in range(template_last_row + 1, template_last_row + needed + 1):
                copy_row_style(ws, style_source_row, row, table.max_col)
            template_last_row += needed
        update_named_ranges_for_capacity(workbook, variant, template_last_row)
        locations = variant_locations(workbook, variant)
        table = locations["gb_data_table"]

        _clear_named_output_range(workbook, locations, "gb_data_table")

        metadata_values = {
            "gb_term": meta["term"],
            "gb_course": meta["course"],
            "gb_teacher": meta["teacher"],
            "gb_class_name": meta["class_name"],
        }
        for name, value in metadata_values.items():
            _named_output_cell(workbook, locations, name).value = value
        _named_output_cell(workbook, locations, "gb_header_regular").value = (
            f"平时成绩({percentage_label(meta['regular_pct'])}%)"
        )
        _named_output_cell(workbook, locations, "gb_header_theory").value = (
            f"理论成绩({percentage_label(meta['theory_pct'])}%)"
        )
        if has_skill:
            _named_output_cell(workbook, locations, "gb_header_skill").value = (
                f"技能成绩（{percentage_label(meta['skill_pct'])}%）"
            )
        else:
            ws.column_dimensions[get_column_letter(locations["gb_total_score_col"].min_col)].width = max(
                ws.column_dimensions[get_column_letter(locations["gb_total_score_col"].min_col)].width or 0,
                18,
            )

        regular_pct = formula_number(meta["regular_pct"])
        theory_pct = formula_number(meta["theory_pct"])
        skill_pct = formula_number(meta["skill_pct"])
        class_code = source_xls.parent.name or source_xls.stem
        regular_start_col = locations["gb_regular_items"].min_col
        regular_end_col = locations["gb_regular_items"].max_col

        for idx, student in enumerate(students):
            row = data_start + idx
            scores = generate_regular_scores(
                student["regular"], f"{class_code}|{student['id']}|{student['regular']}", regular_item_count
            )
            _named_output_cell(workbook, locations, "gb_serial_col", idx).value = idx + 1
            student_id_cell = _named_output_cell(workbook, locations, "gb_student_id_col", idx)
            student_id_cell.value = student["id"]
            student_id_cell.number_format = "@"
            _named_output_cell(workbook, locations, "gb_student_name_col", idx).value = student["name"]
            for offset, score in enumerate(scores):
                _named_output_cell(workbook, locations, "gb_regular_items", idx, offset).value = score
            regular_start = get_column_letter(regular_start_col)
            regular_end = get_column_letter(regular_end_col)
            theory_score = _named_col_letter(locations["gb_theory_score_col"].to_dict())
            _named_output_cell(workbook, locations, "gb_regular_weighted_col", idx).value = (
                f"=AVERAGE({regular_start}{row}:{regular_end}{row})*{regular_pct}"
            )
            _named_output_cell(workbook, locations, "gb_theory_score_col", idx).value = student["theory"]
            _named_output_cell(workbook, locations, "gb_theory_weighted_col", idx).value = (
                f"={theory_score}{row}*{theory_pct}"
            )
            if has_skill:
                skill_score = _named_col_letter(locations["gb_skill_score_col"].to_dict())
                _named_output_cell(workbook, locations, "gb_skill_score_col", idx).value = student["skill"]
                _named_output_cell(workbook, locations, "gb_skill_weighted_col", idx).value = (
                    f"={skill_score}{row}*{skill_pct}"
                )
                _named_output_cell(workbook, locations, "gb_total_score_col", idx).value = (
                    f"=ROUND(AVERAGE({regular_start}{row}:{regular_end}{row})*{regular_pct}+"
                    f"{theory_score}{row}*{theory_pct}+{skill_score}{row}*{skill_pct},0)"
                )
            else:
                _named_output_cell(workbook, locations, "gb_total_score_col", idx).value = (
                    f"=ROUND(AVERAGE({regular_start}{row}:{regular_end}{row})*{regular_pct}+"
                    f"{theory_score}{row}*{theory_pct},0)"
                )

        output_dir.mkdir(parents=True, exist_ok=True)
        temp_xlsx = tmp / f"{class_code}-平时成绩记分册.xlsx"
        workbook.save(temp_xlsx)
        converted = convert_with_soffice(soffice, temp_xlsx, output_dir, "xls")
        final_path = output_dir / f"{class_code}-平时成绩记分册.xls"
        if converted.resolve() != final_path.resolve():
            if final_path.exists():
                final_path.unlink()
            converted.replace(final_path)
        raw_inventory = validate_xls_named_range_inventory(final_path, manifest["anchors"], variant)
        if raw_inventory["errors"]:
            raise NamedRangeError("Generated XLS named-range inventory is invalid: " + "; ".join(raw_inventory["errors"]))
        return {
            "source": str(source_xls),
            "output": str(final_path),
            "count": len(students),
            "course": meta["course"],
            "class_name": meta["class_name"],
            "has_skill": has_skill,
            "named_range_variant": variant,
            "named_range_capacity_end": template_last_row,
            "regular_pct": meta["regular_pct"],
            "theory_pct": meta["theory_pct"],
            "skill_pct": meta["skill_pct"],
            "engine": "libreoffice-openpyxl",
            "platform": platform.system(),
            "normalized_input": normalized,
        }


def build_one(source_xls: Path, template_xls: Path, output_dir: Path, soffice: str, manifest: dict, schema_path: Path) -> dict:
    mode = anchor_mode(manifest)
    if mode == "excel_named_range":
        return build_one_named_ranges(source_xls, template_xls, output_dir, soffice, manifest, schema_path)
    return build_one_legacy(source_xls, template_xls, output_dir, soffice, manifest, schema_path)


def resolve_sources(source_path: Path) -> list[Path]:
    if not source_path.exists():
        raise RuntimeError(f"Source path not found: {source_path}. Provide 课程成绩单.xls or a folder containing it.")
    if source_path.is_dir():
        candidate = source_path / "课程成绩单.xls"
        if not candidate.exists():
            raise RuntimeError(f"Folder does not contain 课程成绩单.xls: {source_path}")
        return [candidate]
    return [source_path]


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate 平时成绩记分册 from 课程成绩单.xls.")
    parser.add_argument("--source", required=True, help="课程成绩单.xls path or a folder containing it")
    parser.add_argument("--output-dir", default="", help="Output directory")
    parser.add_argument("--template", default="", help="Template .xls path")
    parser.add_argument("--manifest", default="", help="Versioned template manifest path")
    parser.add_argument("--schema", default=str(DEFAULT_SCHEMA), help="Normalized input schema path")
    parser.add_argument("--skip-template-validation", action="store_true")
    parser.add_argument("--skip-output-validation", action="store_true")
    parser.add_argument("--qa-report", default="", help="QA report path")
    args = parser.parse_args()

    package = resolve_template_package(args.template or None, args.manifest or None)
    manifest = package.manifest
    ensure_supported_major(manifest)
    validate_manifest_contract(manifest)
    sources = resolve_sources(Path(args.source).expanduser().resolve())
    template = package.template_path
    if not template.exists():
        raise RuntimeError(f"Template not found: {template}")
    validate_template_package_identity(template, manifest)
    output_dir = Path(args.output_dir).expanduser().resolve() if args.output_dir else sources[0].parent / "平时成绩记分册_生成"
    soffice = find_soffice()
    template_warnings: list[str] = []
    if not args.skip_template_validation:
        template_report = validate_template(template, package.manifest_path)
        template_warnings = template_report.get("warnings", [])
        for warning in template_warnings:
            print(f"WARNING: {warning}")
    schema_path = Path(args.schema).expanduser().resolve()
    results = [build_one(source, template, output_dir, soffice, manifest, schema_path) for source in sources]
    if not args.skip_output_validation:
        if len(results) == 1:
            qa_path = Path(args.qa_report).expanduser().resolve() if args.qa_report else None
            report = validate_output_dir(
                output_dir,
                results[0]["normalized_input"],
                manifest,
                qa_path,
                schema_path,
                template_path=template,
                custom_template=bool(args.template),
                engine=results[0]["engine"],
                template_validation=not args.skip_template_validation,
                output_file=results[0]["output"],
                extra_warnings=template_warnings,
            )
            action = "skipped validation" if report["status"] == "skipped" else "validated"
            print(f"{action} files={report['checks']['file_count']['actual']} students={len(report['checks'].get('students', []))} qa={report['qa_report']}")
        else:
            print("WARNING: batch output validation is performed per generated file")
            for result in results:
                output_path = Path(result["output"])
                with tempfile.TemporaryDirectory(prefix="gradebook-validate-") as validation_dir:
                    validation_root = Path(validation_dir)
                    validation_copy = validation_root / output_path.name
                    shutil.copy2(output_path, validation_copy)
                    validate_output_dir(
                        validation_root,
                        result["normalized_input"],
                        manifest,
                        output_dir / f"{output_path.stem}.qa-report.json",
                        schema_path,
                        template_path=template,
                        custom_template=bool(args.template),
                        engine=result["engine"],
                        template_validation=not args.skip_template_validation,
                        output_file=validation_copy,
                        extra_warnings=template_warnings,
                    )
    else:
        if len(results) == 1:
            qa_path = Path(args.qa_report).expanduser().resolve() if args.qa_report else None
            report = write_skipped_report(
                output_dir,
                results[0]["normalized_input"],
                manifest,
                qa_path,
                schema_path,
                template_path=template,
                custom_template=bool(args.template),
                engine=results[0]["engine"],
                template_validation=not args.skip_template_validation,
                output_file=results[0]["output"],
                warnings=template_warnings,
            )
            print(f"WARNING: output validation skipped; qa={report['qa_report']}")
        else:
            for result in results:
                output_path = Path(result["output"])
                report = write_skipped_report(
                    output_dir,
                    result["normalized_input"],
                    manifest,
                    output_dir / f"{output_path.stem}.qa-report.json",
                    schema_path,
                    template_path=template,
                    custom_template=bool(args.template),
                    engine=result["engine"],
                    template_validation=not args.skip_template_validation,
                    output_file=output_path,
                    warnings=template_warnings,
                )
            print("WARNING: output validation skipped; QA reports were written with status=skipped")
    for result in results:
        result.pop("normalized_input", None)
    payload = results[0] if len(results) == 1 else results
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
