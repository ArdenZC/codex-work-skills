from __future__ import annotations

import argparse
import json
import math
import re
import sys
import tempfile
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from package_common import (
    DEFAULT_MANIFEST,
    DEFAULT_SCHEMA,
    calculate_expected_total,
    column_number,
    load_manifest,
    manifest_template_path,
    percentage_label,
    source_total_matches,
    validate_input,
)
from validate_template import (
    _cell_format_signatures_match,
    _cell_format_signature,
    _dimension_signature,
    _non_target_sheet_signature,
    _page_setup_signature,
    _protection_signature,
    _signature_differences,
    convert_to_xlsx,
    find_soffice,
)


def _number(value: Any, label: str) -> float:
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label} is not numeric") from exc


def _cell(ws, column: str, row: int):
    return ws[f"{column}{row}"]


def _formula_number(value: float) -> str:
    return f"{value:.8f}".rstrip("0").rstrip(".") or "0"


def _normalize_formula(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).upper().replace("$", "")


def _contains_formula_error(value: Any) -> bool:
    return isinstance(value, str) and any(
        token in value.upper() for token in ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A")
    )


def _expected_formulas(
    row: int,
    weights: dict[str, Any],
    columns: dict[str, str],
    skill_enabled: bool,
    total_column: str,
) -> dict[str, str]:
    regular = _formula_number(float(weights["regular"]))
    theory = _formula_number(float(weights["theory"]))
    skill = _formula_number(float(weights["skill"]))
    average = f"AVERAGE({columns['regular_items_start']}{row}:{columns['regular_items_end']}{row})"
    expected = {
        columns["regular_weighted"]: f"={average}*{regular}",
        columns["theory_weighted"]: f"={columns['theory_score']}{row}*{theory}",
    }
    total = f"=ROUND({average}*{regular}+{columns['theory_score']}{row}*{theory}"
    if skill_enabled:
        expected[columns["skill_weighted"]] = f"={columns['skill_score']}{row}*{skill}"
        total += f"+{columns['skill_score']}{row}*{skill}"
    expected[total_column] = total + ",0)"
    return expected


def _row_has_content(ws, row: int, max_col: int) -> bool:
    return any(ws.cell(row, col).value not in (None, "") for col in range(1, max_col + 1))


def _non_target_sheet_format_errors(workbook, template_workbook, target_sheet: str) -> list[str]:
    errors: list[str] = []
    for sheet in workbook.worksheets:
        if sheet.title == target_sheet or sheet.title not in template_workbook.sheetnames:
            continue
        expected = _non_target_sheet_signature(template_workbook[sheet.title])
        actual = _non_target_sheet_signature(sheet)
        differences = _signature_differences(expected, actual)
        if differences:
            paths = ",".join(item["path"] for item in differences[:3])
            suffix = f" ({paths})" if paths else ""
            errors.append(f"Protected worksheet changed: {sheet.title}{suffix}")
    return errors


def _check_workbook_protection(
    ws,
    workbook,
    manifest: dict[str, Any],
    skill_enabled: bool,
    errors: list[str],
    template_ws=None,
    template_workbook=None,
) -> dict[str, Any]:
    structure = manifest["structure"]
    validation = manifest["validation"]
    expected_sheets = list(structure.get("required_sheets", []))
    if expected_sheets and workbook.sheetnames != expected_sheets:
        errors.append(f"Worksheet names changed: expected {expected_sheets}, got {workbook.sheetnames}")
    expected_states = {str(key): str(value) for key, value in structure.get("required_sheet_states", {}).items()}
    actual_states = {sheet.title: sheet.sheet_state for sheet in workbook.worksheets}
    if expected_states and actual_states != expected_states:
        errors.append(f"Worksheet visibility changed: expected {expected_states}, got {actual_states}")
    expected_merged_key = "required_merged_ranges" if skill_enabled else "required_merged_ranges_without_skill"
    expected_merged = sorted(str(item).upper() for item in structure.get(expected_merged_key, []))
    actual_merged = sorted(str(item).upper() for item in ws.merged_cells.ranges)
    if expected_merged != actual_merged:
        errors.append(f"Output merged ranges changed: expected {expected_merged}, got {actual_merged}")
    expected_columns = column_number(
        structure["columns"]["total_score"] if skill_enabled else structure["no_skill_total_column"]
    )
    if ws.max_column != expected_columns:
        errors.append(f"Output column count changed: expected {expected_columns}, got {ws.max_column}")
    if ws.page_setup.orientation != validation.get("page_orientation", "landscape"):
        errors.append(f"Output page orientation changed: {ws.page_setup.orientation}")
    expected_print_area = str(validation.get("expected_print_area") or "")
    if "expected_print_area" in validation and str(ws.print_area or "") != expected_print_area:
        errors.append(f"Output print area changed: expected {expected_print_area!r}, got {str(ws.print_area or '')!r}")
    expected_freeze = str(validation.get("expected_freeze_panes") or "")
    if "expected_freeze_panes" in validation and str(ws.freeze_panes or "") != expected_freeze:
        errors.append(f"Output freeze panes changed: expected {expected_freeze!r}, got {str(ws.freeze_panes or '')!r}")
    if template_ws is not None and _page_setup_signature(ws) != _page_setup_signature(template_ws):
        errors.append("target sheet print settings mismatch")
    if template_ws is not None and _protection_signature(ws.protection) != _protection_signature(template_ws.protection):
        errors.append("target sheet protection settings mismatch")
    if template_workbook is not None and _protection_signature(workbook.security) != _protection_signature(template_workbook.security):
        errors.append("workbook protection settings mismatch")
    expected_named_ranges = sorted(str(item) for item in validation.get("required_named_ranges", []))
    actual_named_ranges = sorted(str(name) for name in workbook.defined_names)
    if actual_named_ranges != expected_named_ranges:
        errors.append(f"Output named ranges changed: expected {expected_named_ranges}, got {actual_named_ranges}")
    expected_dv = int(validation.get("required_data_validations", 0))
    actual_dv = len(ws.data_validations.dataValidation)
    if actual_dv != expected_dv:
        errors.append(f"Output data validations changed: expected {expected_dv}, got {actual_dv}")
    expected_cf = int(validation.get("required_conditional_formats", 0))
    actual_cf = len(ws.conditional_formatting)
    if actual_cf != expected_cf:
        errors.append(f"Output conditional formats changed: expected {expected_cf}, got {actual_cf}")
    formatting_errors = _target_sheet_format_errors(ws, template_ws, manifest, skill_enabled) if template_ws is not None else []
    errors.extend(formatting_errors)
    protected_value_errors = (
        _target_protected_value_errors(ws, template_ws, manifest, skill_enabled)
        if template_ws is not None
        else []
    )
    errors.extend(protected_value_errors)
    non_target_errors = (
        _non_target_sheet_format_errors(workbook, template_ws.parent, ws.title)
        if template_ws is not None
        else []
    )
    errors.extend(non_target_errors)
    return {
        "sheets": workbook.sheetnames,
        "sheet_states": actual_states,
        "rows": ws.max_row,
        "columns": ws.max_column,
        "merged_ranges": actual_merged,
        "orientation": ws.page_setup.orientation,
        "print_area": str(ws.print_area or ""),
        "freeze_panes": str(ws.freeze_panes or ""),
        "page_setup": _page_setup_signature(ws),
        "named_ranges": actual_named_ranges,
        "data_validations": actual_dv,
        "conditional_formats": actual_cf,
        "target_formatting_checked": template_ws is not None,
        "target_formatting_error_count": len(formatting_errors),
        "target_protected_values_checked": template_ws is not None,
        "target_protected_value_error_count": len(protected_value_errors),
        "non_target_formatting_checked": template_ws is not None,
        "non_target_formatting_error_count": len(non_target_errors),
    }


def _scan_formula_errors(formula_workbook, value_workbook) -> list[str]:
    errors: list[str] = []
    for sheet in formula_workbook.worksheets:
        value_sheet = value_workbook[sheet.title] if sheet.title in value_workbook.sheetnames else None
        for row in sheet.iter_rows():
            for cell in row:
                if _contains_formula_error(cell.value):
                    errors.append(f"{sheet.title}!{cell.coordinate} contains a formula error: {cell.value}")
                if value_sheet is not None and _contains_formula_error(value_sheet[cell.coordinate].value):
                    errors.append(f"{sheet.title}!{cell.coordinate} has a cached formula error: {value_sheet[cell.coordinate].value}")
    return errors


def _dimension_signature_for_column(sheet, column: int, width_floor: float | None = None) -> tuple[Any, ...]:
    dimension = sheet.column_dimensions.get(get_column_letter(column))
    width = _dimension_signature(dimension.width if dimension is not None else None)
    if width_floor is not None:
        width = max(width or 0, width_floor)
    return (
        width,
        bool(dimension.hidden) if dimension is not None else False,
        int(dimension.outlineLevel) if dimension is not None else 0,
        bool(dimension.collapsed) if dimension is not None else False,
    )


def _dimension_signature_for_row(sheet, row: int) -> tuple[Any, ...]:
    dimension = sheet.row_dimensions.get(row)
    return (
        _dimension_signature(dimension.height if dimension is not None else None),
        bool(dimension.hidden) if dimension is not None else False,
        int(dimension.outlineLevel) if dimension is not None else 0,
        bool(dimension.collapsed) if dimension is not None else False,
    )


def _target_cell_format_signature(cell) -> dict[str, Any]:
    signature = _cell_format_signature(cell)
    alignment = list(signature["alignment"])
    if alignment[0] in (None, "general") or (
        alignment[0] == "left"
        and isinstance(cell.value, str)
        and not cell.value.startswith("=")
    ):
        alignment[0] = "general"
    if alignment[1] in (None, "bottom"):
        alignment[1] = "bottom"
    signature["alignment"] = tuple(alignment)
    return signature


def _format_signature_difference(actual: dict[str, Any], expected: dict[str, Any]) -> str:
    return ",".join(key for key in actual if actual.get(key) != expected.get(key))


def _shift_column_after_delete(column: int, start: int, count: int) -> int | None:
    if column < start:
        return column
    if column >= start + count:
        return column - count
    return None


def _delete_columns_for_signature(ws, start_col: int, count: int) -> None:
    original_ranges = [
        (merged.min_row, merged.max_row, merged.min_col, merged.max_col)
        for merged in list(ws.merged_cells.ranges)
    ]
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    ws.delete_cols(start_col, count)
    for min_row, max_row, min_col, max_col in original_ranges:
        shifted_min = _shift_column_after_delete(min_col, start_col, count)
        shifted_max = _shift_column_after_delete(max_col, start_col, count)
        if shifted_min is None:
            shifted_min = start_col
        if shifted_max is None:
            shifted_max = start_col - 1
        if min_col < start_col and max_col >= start_col + count:
            shifted_min = min_col
            shifted_max = max_col - count
        elif min_col < start_col <= max_col:
            shifted_min = min_col
            shifted_max = start_col - 1
        elif min_col >= start_col and max_col < start_col + count:
            continue
        if shifted_max < shifted_min:
            continue
        if shifted_min != shifted_max or min_row != max_row:
            ws.merge_cells(
                start_row=min_row,
                start_column=shifted_min,
                end_row=max_row,
                end_column=shifted_max,
            )


def _target_sheet_format_errors(output_ws, template_ws, manifest: dict[str, Any], skill_enabled: bool) -> list[str]:
    structure = manifest["structure"]
    columns = structure["columns"]
    template_last_row = int(structure["template_last_data_row"])
    style_source_row = int(structure["style_source_row"])
    output_total_column = column_number(columns["total_score"] if skill_enabled else structure["no_skill_total_column"])
    errors: list[str] = []

    for output_column in range(1, output_total_column + 1):
        expected_dimension = _dimension_signature_for_column(template_ws, output_column)
        actual_dimension = _dimension_signature_for_column(output_ws, output_column)
        if expected_dimension != actual_dimension:
            errors.append(
                f"target sheet formatting mismatch in column {get_column_letter(output_column)}"
            )

    for output_row in range(1, output_ws.max_row + 1):
        source_row = output_row if output_row <= template_last_row else style_source_row
        if source_row > template_ws.max_row:
            errors.append(f"target sheet formatting mismatch in row {output_row}")
            continue
        if _dimension_signature_for_row(output_ws, output_row) != _dimension_signature_for_row(template_ws, source_row):
            errors.append(f"target sheet formatting mismatch in row {output_row}")
        for output_column in range(1, output_total_column + 1):
            if output_column > template_ws.max_column:
                errors.append(
                    f"target sheet formatting mismatch at {get_column_letter(output_column)}{output_row}"
                )
                continue
            actual_signature = _target_cell_format_signature(output_ws.cell(output_row, output_column))
            expected_signature = _target_cell_format_signature(template_ws.cell(source_row, output_column))
            if not _cell_format_signatures_match(actual_signature, expected_signature):
                errors.append(
                    f"target sheet formatting mismatch at {get_column_letter(output_column)}{output_row} "
                    f"({_format_signature_difference(actual_signature, expected_signature)})"
                )
    return errors[:20]


def _shift_cell_address_after_delete(address: str, start_col: int, count: int) -> str | None:
    match = re.fullmatch(r"([A-Z]+)(\d+)", str(address).upper())
    if not match:
        return None
    shifted = _shift_column_after_delete(column_number(match.group(1)), start_col, count)
    if shifted is None:
        return None
    return f"{get_column_letter(shifted)}{match.group(2)}"


def _normalized_cell_value(value: Any) -> Any:
    return value.replace("\r\n", "\n") if isinstance(value, str) else value


def _target_protected_value_errors(output_ws, template_ws, manifest: dict[str, Any], skill_enabled: bool) -> list[str]:
    structure = manifest["structure"]
    writable_cells = {
        str(cell).upper()
        for cell in [*structure.get("metadata", {}).values(), *structure.get("headers", {}).values()]
        if cell
    }
    if not skill_enabled:
        class_name_cell = str(structure.get("metadata", {}).get("class_name", "")).upper()
        shifted_writable_cells: set[str] = set()
        for address in writable_cells:
            if address == class_name_cell:
                shifted_writable_cells.add(address)
                continue
            shifted = _shift_cell_address_after_delete(
                address,
                column_number(structure["columns"]["skill_score"]),
                2,
            )
            if shifted is not None:
                shifted_writable_cells.add(shifted)
        writable_cells = shifted_writable_cells
    data_start_row = int(structure["data_start_row"])
    total_column = column_number(
        structure["columns"]["total_score"] if skill_enabled else structure["no_skill_total_column"]
    )
    errors: list[str] = []
    for row in range(1, data_start_row):
        for column in range(1, total_column + 1):
            address = f"{get_column_letter(column)}{row}"
            if address in writable_cells:
                continue
            expected = _normalized_cell_value(template_ws[address].value)
            actual = _normalized_cell_value(output_ws[address].value)
            if actual != expected:
                errors.append(f"target sheet protected value mismatch at {address}")
    return errors


def _base_qa_report(
    out_dir: Path,
    manifest: dict[str, Any],
    qa_report_path: Path | str | None,
    output_file: Path | None,
    template_path: Path | str | None,
    custom_template: bool | None,
    engine: str | None,
    template_validation: bool,
    output_validation: bool,
    extra_warnings: list[str] | None,
) -> dict[str, Any]:
    canonical_template = manifest_template_path(manifest)
    selected_template = (
        Path(template_path).expanduser().resolve() if template_path else canonical_template
    )
    is_custom_template = (
        bool(custom_template) if custom_template is not None else selected_template != canonical_template
    )
    skipped = []
    if not template_validation:
        skipped.append("template")
    if not output_validation:
        skipped.append("output")
    warnings = list(extra_warnings or [])
    if is_custom_template:
        warnings.append("Custom template selected; output was validated against the supplied manifest.")
    if not template_validation:
        warnings.append("Template validation skipped by explicit flag.")
    if not output_validation:
        warnings.append("Output validation skipped by explicit flag.")
    warnings = list(dict.fromkeys(warnings))
    validation = {"template": template_validation, "output": output_validation, "skipped": skipped}
    report_path = Path(qa_report_path).expanduser().resolve() if qa_report_path else out_dir / "qa-report.json"
    output_label = ""
    if output_file is not None:
        try:
            output_label = output_file.resolve().relative_to(out_dir).as_posix()
        except ValueError:
            output_label = output_file.name
    return {
        "status": "failed",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "template_id": manifest.get("template", {}).get("id"),
        "template_version": manifest.get("template", {}).get("version"),
        "generator_version": manifest.get("generator", {}).get("version"),
        "template_sha256": manifest.get("fingerprint", {}).get("sha256") or manifest.get("fingerprint", {}).get("value"),
        "template_path": str(selected_template),
        "custom_template": is_custom_template,
        "engine": engine or "unknown",
        "validation": validation,
        "validation_skipped": skipped,
        "output_dir": str(out_dir),
        "output_file": output_label,
        "errors": [],
        "warnings": warnings,
        "checks": {"validation": validation},
        "files_checked": 0,
        "qa_report": str(report_path),
    }


def _write_qa_report(report: dict[str, Any]) -> dict[str, Any]:
    report_path = Path(report["qa_report"])
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return report


def write_skipped_report(
    output_dir: Path | str,
    data: dict[str, Any],
    manifest: dict[str, Any] | None = None,
    qa_report_path: Path | str | None = None,
    schema_path: Path | str = DEFAULT_SCHEMA,
    *,
    template_path: Path | str | None = None,
    custom_template: bool | None = None,
    engine: str | None = None,
    template_validation: bool = True,
    output_file: Path | str | None = None,
    warnings: list[str] | None = None,
) -> dict[str, Any]:
    out_dir = Path(output_dir).expanduser().resolve()
    manifest = manifest or load_manifest()
    validate_input(data, schema_path)
    out_dir.mkdir(parents=True, exist_ok=True)
    selected_file = _resolve_output_file(out_dir, output_file)
    files = [selected_file] if selected_file is not None else sorted(out_dir.glob("*.xls"))
    report = _base_qa_report(
        out_dir,
        manifest,
        qa_report_path,
        selected_file,
        template_path,
        custom_template,
        engine,
        template_validation,
        False,
        warnings,
    )
    report["status"] = "skipped"
    report["checks"]["file_count"] = {"expected": 1, "actual": len(files)}
    report["files_checked"] = len(files)
    return _write_qa_report(report)


def _resolve_output_file(out_dir: Path, output_file: Path | str | None) -> Path | None:
    if output_file is None or str(output_file).strip() == "":
        return None
    candidate = Path(output_file).expanduser()
    if not candidate.is_absolute():
        candidate = out_dir / candidate
    candidate = candidate.resolve()
    try:
        candidate.relative_to(out_dir)
    except ValueError as exc:
        raise ValueError("--output-file must be inside --output-dir") from exc
    return candidate


def validate_output_dir(
    output_dir: Path | str,
    data: dict[str, Any],
    manifest: dict[str, Any] | None = None,
    qa_report_path: Path | str | None = None,
    schema_path: Path | str = DEFAULT_SCHEMA,
    *,
    template_path: Path | str | None = None,
    custom_template: bool | None = None,
    engine: str | None = None,
    template_validation: bool = True,
    output_validation: bool = True,
    output_file: Path | str | None = None,
    extra_warnings: list[str] | None = None,
) -> dict[str, Any]:
    out_dir = Path(output_dir).expanduser().resolve()
    manifest = manifest or load_manifest()
    validate_input(data, schema_path)
    selected_file = _resolve_output_file(out_dir, output_file)
    files = [selected_file] if selected_file is not None else sorted(out_dir.glob("*.xls"))
    report = _base_qa_report(
        out_dir,
        manifest,
        qa_report_path,
        selected_file,
        template_path,
        custom_template,
        engine,
        template_validation,
        output_validation,
        extra_warnings,
    )
    errors: list[str] = report["errors"]
    if selected_file is None and len(files) != 1:
        errors.append(f"Expected one generated XLS file, got {len(files)}")
    path = None
    if selected_file is not None:
        if not selected_file.exists():
            errors.append(f"Generated XLS file not found: {selected_file.name}")
        elif not selected_file.is_file():
            errors.append(f"Generated XLS path is not a file: {selected_file.name}")
        else:
            path = selected_file
    elif len(files) == 1:
        path = files[0]
    if path is not None:
        structure = manifest["structure"]
        columns = structure["columns"]
        start_row = int(structure["data_start_row"])
        skill_enabled = float(data["weights"]["skill"]) > 0.000001
        total_column = columns["total_score"] if skill_enabled else structure["no_skill_total_column"]
        formula_columns = (
            manifest["fields"]["formula_columns_with_skill"]["columns"]
            if skill_enabled
            else manifest["fields"]["formula_columns_without_skill"]["columns"]
        )
        try:
            if path.stat().st_size == 0:
                raise RuntimeError("Generated XLS file is empty")
            with tempfile.TemporaryDirectory(prefix="gradebook-output-") as temp_name:
                xlsx = convert_to_xlsx(path, Path(temp_name), find_soffice())
                formulas = load_workbook(xlsx, data_only=False)
                values = load_workbook(xlsx, data_only=True)
                selected_template = Path(template_path).expanduser().resolve() if template_path else manifest_template_path(manifest)
                template_xlsx = (
                    selected_template
                    if selected_template.suffix.lower() == ".xlsx"
                    else convert_to_xlsx(selected_template, Path(temp_name) / "template", find_soffice())
                )
                template_workbook = load_workbook(template_xlsx, data_only=False)
                sheet_name = structure["worksheet"]
                if sheet_name not in formulas.sheetnames:
                    errors.append(f"Missing worksheet: {sheet_name}")
                else:
                    ws_formula = formulas[sheet_name]
                    ws_values = values[sheet_name]
                    template_ws = template_workbook[sheet_name] if sheet_name in template_workbook.sheetnames else None
                    if template_ws is None:
                        errors.append(f"Missing worksheet in template: {sheet_name}")
                    elif not skill_enabled:
                        class_name_cell = structure["metadata"]["class_name"]
                        class_name_style = copy(template_ws[class_name_cell]._style)
                        _delete_columns_for_signature(
                            template_ws,
                            column_number(columns["skill_score"]),
                            2,
                        )
                        template_ws[class_name_cell]._style = copy(class_name_style)
                        no_skill_total = structure["no_skill_total_column"]
                        template_ws.column_dimensions[no_skill_total].width = max(
                            template_ws.column_dimensions[no_skill_total].width or 0,
                            18,
                        )
                    report["checks"]["structure"] = _check_workbook_protection(
                        ws_formula,
                        formulas,
                        manifest,
                        skill_enabled,
                        errors,
                        template_ws,
                        template_workbook,
                    )
                    errors.extend(_scan_formula_errors(formulas, values))

                    expected_headers = {
                        "regular": f"平时成绩({percentage_label(data['weights']['regular'])}%)",
                        "theory": f"理论成绩({percentage_label(data['weights']['theory'])}%)",
                        "total": "总评\n成绩",
                    }
                    header_cells = structure["headers"]
                    for name in ("regular", "theory"):
                        actual = str(ws_values[header_cells[name]].value or "").replace("\r\n", "\n")
                        if actual != expected_headers[name]:
                            errors.append(f"{name} header mismatch: expected {expected_headers[name]!r}, got {actual!r}")
                    total_header = columns["total_score"] if skill_enabled else structure["no_skill_total_column"]
                    actual_total_header = str(ws_values[f"{total_header}3"].value or "").replace("\r\n", "\n")
                    if actual_total_header != expected_headers["total"]:
                        errors.append(f"total header mismatch: expected {expected_headers['total']!r}, got {actual_total_header!r}")
                    if skill_enabled:
                        skill_header = str(ws_values[header_cells["skill"]].value or "").replace("\r\n", "\n")
                        expected_skill_header = f"技能成绩（{percentage_label(data['weights']['skill'])}%）"
                        if skill_header != expected_skill_header:
                            errors.append(f"skill header mismatch: expected {expected_skill_header!r}, got {skill_header!r}")
                    else:
                        for cell in ("O3", "P3", "Q3", "O4", "P4", "Q4"):
                            if "技能成绩" in str(ws_values[cell].value or ""):
                                errors.append(f"skill header remains after zero-skill column removal: {cell}")

                    expected_meta = {
                        "term": data["term"],
                        "course": data["course"],
                        "teacher": data["teacher"],
                        "class_name": data["class_name"],
                    }
                    for name, expected in expected_meta.items():
                        cell = structure["metadata"][name]
                        actual = ws_values[cell].value
                        if str(actual).strip() != str(expected).strip():
                            errors.append(f"{name} metadata mismatch")

                    expected_last_row = start_row + len(data["students"]) - 1
                    if ws_values.max_row != expected_last_row:
                        errors.append(f"Output student row extent mismatch: expected through row {expected_last_row}, got {ws_values.max_row}")
                    output_max_col = column_number(total_column)
                    for extra_row in range(expected_last_row + 1, max(ws_values.max_row, expected_last_row) + 1):
                        if _row_has_content(ws_values, extra_row, output_max_col):
                            errors.append(f"Unexpected non-empty student row after expected data: {extra_row}")

                    student_checks = []
                    regular_start = column_number(columns["regular_items_start"])
                    regular_end = column_number(columns["regular_items_end"])
                    for index, student in enumerate(data["students"]):
                        row = start_row + index
                        row_errors: list[str] = []
                        actual_id = _cell(ws_values, columns["student_id"], row).value
                        if str(actual_id).strip() != student["id"]:
                            row_errors.append("student ID mismatch")
                        if _cell(ws_formula, columns["student_id"], row).number_format != "@":
                            row_errors.append("student ID is not text-formatted")
                        actual_name = _cell(ws_values, columns["student_name"], row).value
                        if str(actual_name).strip() != student["name"]:
                            row_errors.append("student name mismatch")

                        regular_values = [
                            _number(ws_values.cell(row, col).value, f"regular score {row}")
                            for col in range(regular_start, regular_end + 1)
                        ]
                        if len(regular_values) != int(manifest["validation"]["regular_item_count"]):
                            row_errors.append("regular score count changed")
                        for score in regular_values:
                            if not 0 <= score <= 100 or not math.isclose(score * 2, round(score * 2), abs_tol=0.000001):
                                row_errors.append("regular score is outside 0..100 or not a half-point")
                        average = sum(regular_values) / len(regular_values)
                        if not math.isclose(average, float(student["regular"]), abs_tol=0.001):
                            row_errors.append("regular score average mismatch")

                        theory_value = _number(_cell(ws_values, columns["theory_score"], row).value, f"theory score {row}")
                        if not 0 <= theory_value <= 100:
                            row_errors.append("theory score outside 0..100")
                        if not math.isclose(theory_value, float(student["theory"]), abs_tol=0.001):
                            row_errors.append("theory score mismatch")
                        expected_formulas = _expected_formulas(row, data["weights"], columns, skill_enabled, total_column)
                        for col in formula_columns:
                            formula = _cell(ws_formula, col, row).value
                            expected_formula = expected_formulas.get(col)
                            if not isinstance(formula, str) or not formula.startswith("=") or _contains_formula_error(formula):
                                row_errors.append(f"formula missing or broken in {col}{row}")
                            elif expected_formula and _normalize_formula(formula) != _normalize_formula(expected_formula):
                                row_errors.append(f"formula mismatch in {col}{row}: expected {expected_formula}, got {formula}")
                            cached_formula_value = _cell(ws_values, col, row).value
                            if cached_formula_value is None or _contains_formula_error(cached_formula_value):
                                row_errors.append(f"formula result missing or broken in {col}{row}")

                        try:
                            total_number = _number(_cell(ws_values, total_column, row).value, f"total {row}")
                            skill_value = 0.0
                            if skill_enabled:
                                skill_value = _number(_cell(ws_values, columns["skill_score"], row).value, f"skill score {row}")
                                if not 0 <= skill_value <= 100:
                                    row_errors.append("skill score outside 0..100")
                                if not math.isclose(skill_value, float(student["skill"]), abs_tol=0.001):
                                    row_errors.append("skill score mismatch")
                            output_total = calculate_expected_total(
                                {"regular": student["regular"], "theory": theory_value, "skill": skill_value},
                                data["weights"],
                            )
                            source_total = calculate_expected_total(student, data["weights"])
                            if total_number != output_total:
                                row_errors.append("total formula result mismatch")
                            if not source_total_matches(student["total"], source_total):
                                row_errors.append("total differs from source")
                        except ValueError as exc:
                            row_errors.append(str(exc))

                        if row_errors:
                            errors.extend(f"generated workbook row {row}: {message}" for message in row_errors)
                        student_checks.append({"row": row, "status": "failed" if row_errors else "passed", "error_count": len(row_errors)})
                    report["checks"]["students"] = student_checks
                    report["checks"]["skill_enabled"] = skill_enabled
                    report["checks"]["formula_columns"] = formula_columns
        except Exception as exc:
            errors.append(f"Generated XLS could not be opened or inspected: {exc}")
    report["checks"]["file_count"] = {"expected": 1, "actual": len(files)}
    report["files_checked"] = 1 if path is not None else 0
    if not errors:
        report["status"] = "skipped" if report["validation_skipped"] else "passed"
    _write_qa_report(report)
    if errors:
        raise RuntimeError("Output validation failed: " + "; ".join(errors[:8]))
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate a generated XLS gradebook and write a QA report.")
    parser.add_argument("--input-json", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--output-file", default="")
    parser.add_argument("--manifest", default=str(DEFAULT_MANIFEST))
    parser.add_argument("--schema", default=str(DEFAULT_SCHEMA))
    parser.add_argument("--qa-report", default="")
    parser.add_argument("--template-path", default="")
    parser.add_argument("--custom-template", action="store_true")
    parser.add_argument("--engine", default="")
    parser.add_argument("--skip-template-validation", action="store_true")
    parser.add_argument("--skip-validation", action="store_true")
    args = parser.parse_args()
    try:
        data = json.loads(Path(args.input_json).read_text(encoding="utf-8-sig"))
        manifest = load_manifest(args.manifest)
        template_path = args.template_path or None
        custom_template = args.custom_template if args.custom_template else None
        if args.skip_validation:
            report = write_skipped_report(
                args.output_dir,
                data,
                manifest,
                args.qa_report or None,
                args.schema,
                template_path=template_path,
                custom_template=custom_template,
                engine=args.engine or None,
                template_validation=not args.skip_template_validation,
                output_file=args.output_file or None,
            )
        else:
            report = validate_output_dir(
                args.output_dir,
                data,
                manifest,
                args.qa_report or None,
                args.schema,
                template_path=template_path,
                custom_template=custom_template,
                engine=args.engine or None,
                template_validation=not args.skip_template_validation,
                output_file=args.output_file or None,
            )
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    action = "skipped validation" if report["status"] == "skipped" else "validated"
    print(f"{action} files={report['checks']['file_count']['actual']} students={len(report['checks'].get('students', []))} qa={report['qa_report']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
