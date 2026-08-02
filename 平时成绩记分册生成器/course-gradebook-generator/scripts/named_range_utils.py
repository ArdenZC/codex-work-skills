"""OpenPyXL helpers for workbook-level managed gradebook names."""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass, replace
from typing import Any, Iterable

from openpyxl.utils import column_index_from_string, get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName

from named_range_contracts import (
    MANAGED_NAME_PREFIX,
    NAMED_RANGE_CONTRACTS,
    removed_names,
    required_names,
)


_NAME_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]{0,63}$")
_A1_RE = re.compile(
    r"^\$?([A-Za-z]{1,3})\$?(\d+)(?::\$?([A-Za-z]{1,3})\$?(\d+))?$"
)
_BUILTIN_NAMES = {"print_area", "print_titles", "_filterdatabase"}


class NamedRangeError(ValueError):
    def __init__(self, message: str, code: str = "broken") -> None:
        self.code = code
        super().__init__(message)


@dataclass(frozen=True)
class NamedRangeLocation:
    name: str
    scope: str
    sheet: str
    min_row: int
    min_col: int
    max_row: int
    max_col: int

    @property
    def width(self) -> int:
        return self.max_col - self.min_col + 1

    @property
    def height(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def address(self) -> str:
        start = f"${get_column_letter(self.min_col)}${self.min_row}"
        end = f"${get_column_letter(self.max_col)}${self.max_row}"
        return start if start == end else f"{start}:{end}"

    def to_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "scope": self.scope,
            "sheet": self.sheet,
            "min_row": self.min_row,
            "min_col": self.min_col,
            "max_row": self.max_row,
            "max_col": self.max_col,
            "address": self.address,
        }


def _defined_names(workbook) -> list[DefinedName]:
    return [value for value in workbook.defined_names.values() if isinstance(value, DefinedName)]


def _managed_defined_names(workbook) -> list[DefinedName]:
    return [item for item in _defined_names(workbook) if str(item.name).lower().startswith(MANAGED_NAME_PREFIX)]


def _unquote_sheet(value: str) -> str:
    value = value.strip()
    if value.startswith("'") and value.endswith("'"):
        return value[1:-1].replace("''", "'")
    return value


def _split_reference(attr_text: str) -> tuple[str, str]:
    reference = str(attr_text or "").strip()
    if reference.startswith("="):
        reference = reference[1:].strip()
    if not reference or "!" not in reference:
        raise NamedRangeError(f"Named range destination is not a worksheet range: {attr_text!r}", "broken")
    if any(token in reference for token in ("#REF!", "[", "]", ",")):
        raise NamedRangeError(f"Named range destination is broken or non-contiguous: {attr_text!r}", "broken")
    sheet_part, address = reference.rsplit("!", 1)
    if "!" in address:
        raise NamedRangeError(f"Named range destination contains multiple sheets: {attr_text!r}", "destination")
    return _unquote_sheet(sheet_part), address


def _location_from_defined_name(workbook, defined: DefinedName) -> NamedRangeLocation:
    name = str(defined.name)
    if not _NAME_RE.fullmatch(name) or not name.startswith(MANAGED_NAME_PREFIX):
        raise NamedRangeError(f"Invalid managed named range name: {name}", "name")
    if name.lower() in _BUILTIN_NAMES:
        raise NamedRangeError(f"Managed name uses a built-in name: {name}", "name")
    if bool(defined.hidden):
        raise NamedRangeError(f"Managed named range must not be hidden: {name}", "name")
    if defined.localSheetId is not None:
        raise NamedRangeError(f"Managed named range must be workbook scoped: {name}", "scope")
    sheet, address = _split_reference(defined.attr_text)
    match = _A1_RE.fullmatch(address)
    if match is None:
        raise NamedRangeError(f"Named range {name} has an invalid A1 destination: {address!r}", "shape")
    min_col = column_index_from_string(match.group(1))
    min_row = int(match.group(2))
    max_col = column_index_from_string(match.group(3) or match.group(1))
    max_row = int(match.group(4) or match.group(2))
    if min_row < 1 or max_row < min_row or max_col < min_col:
        raise NamedRangeError(f"Named range {name} has an invalid destination: {address!r}", "shape")
    if sheet not in workbook.sheetnames:
        raise NamedRangeError(f"Named range {name} targets missing worksheet {sheet!r}", "destination")
    location = NamedRangeLocation(name, "workbook", sheet, min_row, min_col, max_row, max_col)
    if location.width == 1 and location.height == 1:
        for merged in workbook[sheet].merged_cells.ranges:
            if (
                merged.min_row <= min_row <= merged.max_row
                and merged.min_col <= min_col <= merged.max_col
            ):
                if (min_row, min_col) != (merged.min_row, merged.min_col):
                    raise NamedRangeError(
                        f"Named range {name} must target the top-left cell of merged range {merged}",
                        "shape",
                    )
                break
    return location


def resolve_named_range(workbook, name: str) -> NamedRangeLocation:
    matches = [item for item in _defined_names(workbook) if str(item.name) == str(name)]
    if not matches:
        raise NamedRangeError(f"Missing managed named range: {name}", "missing")
    if len(matches) != 1:
        raise NamedRangeError(f"Duplicate managed named range: {name}", "duplicate")
    return _location_from_defined_name(workbook, matches[0])


def resolve_named_cell(workbook, name: str) -> NamedRangeLocation:
    location = resolve_named_range(workbook, name)
    if location.width != 1 or location.height != 1:
        raise NamedRangeError(f"Named range {name} must target one cell, got {location.address}", "shape")
    return location


def resolve_named_rect(workbook, name: str) -> NamedRangeLocation:
    return resolve_named_range(workbook, name)


def list_managed_names(workbook) -> list[NamedRangeLocation]:
    return sorted(
        (_location_from_defined_name(workbook, item) for item in _managed_defined_names(workbook)),
        key=lambda item: item.name,
    )


def _add_unique(target: list[str], value: str) -> None:
    if value not in target:
        target.append(value)


def _contract_for(contract: dict[str, Any] | None) -> dict[str, dict[str, Any]]:
    if not contract:
        return {name: dict(value) for name, value in NAMED_RANGE_CONTRACTS.items()}
    definitions = contract.get("definitions", contract)
    if isinstance(definitions, list):
        return {str(name): dict(NAMED_RANGE_CONTRACTS[str(name)]) for name in definitions}
    return {str(name): dict(value) for name, value in definitions.items()}


def _column_number(column: str) -> int:
    value = 0
    for char in str(column).upper():
        if not "A" <= char <= "Z":
            raise ValueError(f"Invalid Excel column: {column}")
        value = value * 26 + ord(char) - ord("A") + 1
    return value


def expected_named_range_locations(
    legacy_structure: dict[str, Any],
    variant: str,
) -> dict[str, NamedRangeLocation]:
    """Derive the v1.1 initial layout from the protected v1.0 coordinate contract."""
    sheet = str(legacy_structure["worksheet"])
    start_row = int(legacy_structure["data_start_row"])
    end_row = int(legacy_structure["template_last_data_row"])
    columns = legacy_structure["columns"]

    def cell(name: str, address: str) -> NamedRangeLocation:
        match = re.fullmatch(r"([A-Z]+)(\d+)", str(address).upper())
        if match is None:
            raise ValueError(f"Invalid legacy cell address for {name}: {address}")
        column = _column_number(match.group(1))
        row = int(match.group(2))
        return NamedRangeLocation(name, "workbook", sheet, row, column, row, column)

    def rect(name: str, min_col: str, max_col: str, min_row: int = start_row, max_row: int = end_row) -> NamedRangeLocation:
        return NamedRangeLocation(
            name,
            "workbook",
            sheet,
            min_row,
            _column_number(min_col),
            max_row,
            _column_number(max_col),
        )

    locations = {
        "gb_title": cell("gb_title", legacy_structure["title_cell"]),
        "gb_term": cell("gb_term", legacy_structure["metadata"]["term"]),
        "gb_course": cell("gb_course", legacy_structure["metadata"]["course"]),
        "gb_teacher": cell("gb_teacher", legacy_structure["metadata"]["teacher"]),
        "gb_class_name": cell("gb_class_name", legacy_structure["metadata"]["class_name"]),
        "gb_header_serial": cell("gb_header_serial", legacy_structure["header_label_cells"]["serial"]),
        "gb_header_student_id": cell("gb_header_student_id", legacy_structure["header_label_cells"]["student_id"]),
        "gb_header_student_name": cell("gb_header_student_name", legacy_structure["header_label_cells"]["student_name"]),
        "gb_header_regular": cell("gb_header_regular", legacy_structure["header_label_cells"]["regular"]),
        "gb_header_theory": cell("gb_header_theory", legacy_structure["header_label_cells"]["theory"]),
        "gb_header_total": cell("gb_header_total", legacy_structure["header_label_cells"]["total"]),
        "gb_data_table": rect("gb_data_table", columns["serial"], columns["total_score"]),
        "gb_template_row": rect(
            "gb_template_row",
            columns["serial"],
            columns["total_score"],
            start_row,
            start_row,
        ),
        "gb_serial_col": rect("gb_serial_col", columns["serial"], columns["serial"]),
        "gb_student_id_col": rect("gb_student_id_col", columns["student_id"], columns["student_id"]),
        "gb_student_name_col": rect("gb_student_name_col", columns["student_name"], columns["student_name"]),
        "gb_regular_items": rect("gb_regular_items", columns["regular_items_start"], columns["regular_items_end"]),
        "gb_regular_weighted_col": rect("gb_regular_weighted_col", columns["regular_weighted"], columns["regular_weighted"]),
        "gb_theory_score_col": rect("gb_theory_score_col", columns["theory_score"], columns["theory_score"]),
        "gb_theory_weighted_col": rect("gb_theory_weighted_col", columns["theory_weighted"], columns["theory_weighted"]),
        "gb_skill_score_col": rect("gb_skill_score_col", columns["skill_score"], columns["skill_score"]),
        "gb_skill_weighted_col": rect("gb_skill_weighted_col", columns["skill_weighted"], columns["skill_weighted"]),
        "gb_total_score_col": rect("gb_total_score_col", columns["total_score"], columns["total_score"]),
    }
    if variant == "without_skill":
        total_column = str(legacy_structure["no_skill_total_column"])
        locations["gb_header_total"] = rect("gb_header_total", total_column, total_column, 3, 3)
        locations["gb_data_table"] = rect("gb_data_table", columns["serial"], total_column)
        locations["gb_template_row"] = rect("gb_template_row", columns["serial"], total_column, start_row, start_row)
        locations["gb_total_score_col"] = rect("gb_total_score_col", total_column, total_column)
        for name in ("gb_header_skill", "gb_skill_score_col", "gb_skill_weighted_col"):
            locations.pop(name, None)
    else:
        locations["gb_header_skill"] = cell("gb_header_skill", legacy_structure["headers"]["skill"])
    return {name: locations[name] for name in required_names(variant)}


def _expected_for_variant(contract: dict[str, Any] | None, variant: str) -> tuple[str, ...]:
    if contract and "variants" in contract:
        variant_data = contract["variants"].get(variant, {})
        return tuple(str(name) for name in variant_data.get("required", required_names(variant)))
    return required_names(variant)


def _validate_shapes(
    locations: dict[str, NamedRangeLocation],
    definitions: dict[str, dict[str, Any]],
    expected_names: Iterable[str],
    shape_errors: list[str],
) -> None:
    for name in expected_names:
        location = locations.get(name)
        if location is None:
            continue
        definition = definitions.get(name, NAMED_RANGE_CONTRACTS.get(name, {}))
        kind = definition.get("kind")
        if kind == "cell" and (location.width != 1 or location.height != 1):
            _add_unique(shape_errors, f"{name} must be a single cell, got {location.address}")
        elif kind == "row" and location.height != 1:
            _add_unique(shape_errors, f"{name} must be one row, got {location.address}")
        elif kind == "column" and location.width != 1:
            _add_unique(shape_errors, f"{name} must be one column, got {location.address}")
        elif kind == "matrix" and (location.width < 1 or location.height < 1):
            _add_unique(shape_errors, f"{name} must be a non-empty rectangle, got {location.address}")
        if definition.get("columns") is not None and location.width != int(definition["columns"]):
            _add_unique(shape_errors, f"{name} must be {definition['columns']} columns, got {location.width}")


def _validate_relationships(locations: dict[str, NamedRangeLocation], errors: list[str]) -> None:
    if not locations:
        return
    sheets = {location.sheet for location in locations.values()}
    if len(sheets) != 1:
        _add_unique(errors, f"Managed named ranges must target one worksheet, got {sorted(sheets)}")
    table = locations.get("gb_data_table")
    template_row = locations.get("gb_template_row")
    if table and template_row:
        if (
            template_row.min_row != table.min_row
            or template_row.min_col != table.min_col
            or template_row.max_col != table.max_col
        ):
            _add_unique(errors, "Named range gb_template_row must equal the first row of gb_data_table")
    data_columns = (
        "gb_serial_col",
        "gb_student_id_col",
        "gb_student_name_col",
        "gb_regular_items",
        "gb_regular_weighted_col",
        "gb_theory_score_col",
        "gb_theory_weighted_col",
        "gb_skill_score_col",
        "gb_skill_weighted_col",
        "gb_total_score_col",
    )
    if table:
        for name in data_columns:
            location = locations.get(name)
            if location is None:
                continue
            if location.min_row != table.min_row or location.max_row != table.max_row:
                _add_unique(errors, f"Named range {name} must use the same data rows as gb_data_table")
            if location.min_col < table.min_col or location.max_col > table.max_col:
                _add_unique(errors, f"Named range {name} must be contained in gb_data_table")
        regular = locations.get("gb_regular_items")
        weighted = locations.get("gb_regular_weighted_col")
        if regular and weighted and regular.max_col + 1 != weighted.min_col:
            _add_unique(errors, "Named range gb_regular_items must be immediately before gb_regular_weighted_col")
    physical = Counter(
        (location.sheet, location.min_row, location.min_col, location.max_row, location.max_col)
        for location in locations.values()
    )
    for key, count in physical.items():
        if count > 1:
            names = sorted(
                location.name
                for location in locations.values()
                if (location.sheet, location.min_row, location.min_col, location.max_row, location.max_col) == key
            )
            _add_unique(errors, f"Managed names share one physical destination: {', '.join(names)}")


def validate_named_range_inventory(
    workbook,
    contract: dict[str, Any] | None,
    variant: str,
) -> dict[str, Any]:
    """Return a stable, diagnostic inventory without silently accepting bad names."""
    definitions = _contract_for(contract)
    expected_names = tuple(_expected_for_variant(contract, variant))
    report: dict[str, Any] = {
        "required": sorted(expected_names),
        "required_count": len(expected_names),
        "actual_count": 0,
        "locations": {},
        "missing": [],
        "duplicate": [],
        "invalid_names": [],
        "unexpected": [],
        "scope_errors": [],
        "broken": [],
        "destination_errors": [],
        "shape_errors": [],
        "relationship_errors": [],
        "errors": [],
    }
    records = _managed_defined_names(workbook)
    counts = Counter(str(item.name) for item in records)
    for name, count in sorted(counts.items()):
        if count > 1:
            _add_unique(report["duplicate"], name)
    expected_set = set(expected_names)
    for name in sorted(counts):
        if name not in expected_set:
            _add_unique(report["unexpected"], name)
    locations: dict[str, NamedRangeLocation] = {}
    for defined in sorted(records, key=lambda item: str(item.name)):
        name = str(defined.name)
        if name not in expected_set or counts[name] != 1:
            continue
        try:
            locations[name] = _location_from_defined_name(workbook, defined)
        except NamedRangeError as exc:
            if exc.code == "name":
                _add_unique(report["invalid_names"], name)
            elif exc.code == "scope":
                _add_unique(report["scope_errors"], str(exc))
            elif exc.code == "destination":
                _add_unique(report["destination_errors"], str(exc))
            elif exc.code == "shape":
                _add_unique(report["shape_errors"], str(exc))
            else:
                _add_unique(report["broken"], str(exc))
    for name in expected_names:
        if name not in counts:
            _add_unique(report["missing"], name)
    _validate_shapes(locations, definitions, expected_names, report["shape_errors"])
    _validate_relationships(locations, report["relationship_errors"])
    report["actual_count"] = len(locations)
    report["locations"] = {name: locations[name].to_dict() for name in sorted(locations)}
    report["errors"] = sorted(
        set(
            [f"Missing managed named range: {name}" for name in report["missing"]]
            + [f"Duplicate managed named range: {name}" for name in report["duplicate"]]
            + [f"Invalid managed named range: {name}" for name in report["invalid_names"]]
            + [f"Unexpected managed named range: {name}" for name in report["unexpected"]]
            + report["scope_errors"]
            + report["broken"]
            + report["destination_errors"]
            + report["shape_errors"]
            + report["relationship_errors"]
        )
    )
    return report


def compare_named_range_inventories(
    expected: dict[str, Any],
    actual: dict[str, Any],
    names: Iterable[str] | None = None,
) -> list[str]:
    expected_locations = expected.get("locations", expected)
    actual_locations = actual.get("locations", actual)
    selected = sorted(names or set(expected_locations) | set(actual_locations))
    differences: list[str] = []
    for name in selected:
        left = expected_locations.get(name)
        right = actual_locations.get(name)
        if left is None or right is None:
            continue
        left_key = tuple(left[key] for key in ("scope", "sheet", "min_row", "min_col", "max_row", "max_col"))
        right_key = tuple(right[key] for key in ("scope", "sheet", "min_row", "min_col", "max_row", "max_col"))
        if left_key != right_key:
            differences.append(
                f"Named range destination mismatch for {name}: expected {left.get('sheet')}!{left.get('address')}, "
                f"got {right.get('sheet')}!{right.get('address')}"
            )
    return sorted(set(differences))


def _defined_name_reference(location: NamedRangeLocation) -> str:
    return f"{quote_sheetname(location.sheet)}!{location.address}"


def remove_named_range(workbook, name: str) -> None:
    for key in list(workbook.defined_names):
        defined = workbook.defined_names[key]
        if str(defined.name) == str(name):
            del workbook.defined_names[key]


def set_named_range(
    workbook,
    name: str,
    sheet: str,
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
) -> NamedRangeLocation:
    location = NamedRangeLocation(name, "workbook", sheet, min_row, min_col, max_row, max_col)
    if not _NAME_RE.fullmatch(name) or not name.startswith(MANAGED_NAME_PREFIX):
        raise NamedRangeError(f"Invalid managed named range name: {name}", "name")
    if sheet not in workbook.sheetnames:
        raise NamedRangeError(f"Cannot create {name}: worksheet {sheet!r} does not exist", "destination")
    remove_named_range(workbook, name)
    workbook.defined_names.add(
        DefinedName(name=name, hidden=False, localSheetId=None, attr_text=_defined_name_reference(location))
    )
    return location


def set_named_range_from_location(workbook, location: NamedRangeLocation) -> NamedRangeLocation:
    return set_named_range(
        workbook,
        location.name,
        location.sheet,
        location.min_row,
        location.min_col,
        location.max_row,
        location.max_col,
    )


def shift_location_after_column_delete(
    location: NamedRangeLocation,
    start_col: int,
    count: int,
) -> NamedRangeLocation | None:
    end_col = start_col + count - 1
    if location.max_col < start_col:
        return location
    if location.min_col > end_col:
        return replace(location, min_col=location.min_col - count, max_col=location.max_col - count)
    if location.min_col >= start_col and location.max_col <= end_col:
        return replace(location, min_col=start_col, max_col=start_col)
    if location.min_col < start_col and location.max_col > end_col:
        return replace(location, max_col=location.max_col - count)
    if location.min_col < start_col <= location.max_col <= end_col:
        return replace(location, max_col=start_col - 1)
    if start_col <= location.min_col <= end_col < location.max_col:
        return replace(location, min_col=start_col, max_col=location.max_col - count)
    return None


def rebuild_named_ranges_after_column_delete(
    workbook,
    start_col: int,
    count: int,
    variant: str,
) -> dict[str, NamedRangeLocation]:
    current = {location.name: location for location in list_managed_names(workbook)}
    required = set(required_names(variant))
    for name in set(current) - required:
        remove_named_range(workbook, name)
    shifted: dict[str, NamedRangeLocation] = {}
    for name in sorted(required & set(current)):
        location = shift_location_after_column_delete(current[name], start_col, count)
        if location is None:
            raise NamedRangeError(f"Could not shift managed named range {name} after column deletion")
        shifted[name] = set_named_range_from_location(workbook, location)
    return shifted


def update_named_ranges_for_capacity(
    workbook,
    variant: str,
    last_row: int,
) -> dict[str, NamedRangeLocation]:
    data_range_names = {
        "gb_data_table",
        "gb_serial_col",
        "gb_student_id_col",
        "gb_student_name_col",
        "gb_regular_items",
        "gb_regular_weighted_col",
        "gb_theory_score_col",
        "gb_theory_weighted_col",
        "gb_skill_score_col",
        "gb_skill_weighted_col",
        "gb_total_score_col",
    }
    locations = {location.name: location for location in list_managed_names(workbook)}
    for name in required_names(variant):
        if name not in locations:
            raise NamedRangeError(f"Cannot update missing managed named range {name}")
        location = locations[name]
        if name in data_range_names:
            location = replace(location, max_row=last_row)
        locations[name] = set_named_range_from_location(workbook, location)
    return locations


def variant_locations(workbook, variant: str) -> dict[str, NamedRangeLocation]:
    expected = set(required_names(variant))
    return {location.name: location for location in list_managed_names(workbook) if location.name in expected}
